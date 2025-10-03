#!/usr/bin/env python3
"""
Analyseur de Trading Unifié
Combine l'analyse Forex et autres instruments dans un seul script
"""

import pandas as pd
import os
import re
import math
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, PieChart, BarChart
from openpyxl.chart.label import DataLabelList
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum

class InstrumentType(Enum):
    FOREX = "forex"
    METAUX = "metaux"
    INDICES = "indices"
    CRYPTO = "crypto"
    ENERGIE = "energie"
    ACTIONS = "actions"

class TradingAnalyzer:
    def __init__(self, solde_initial=10000):
        self.solde_initial = solde_initial
        self.statistiques_fichiers = {}
        
        # Configuration des symboles par type
        self.symboles_forex = [
            "eurusd", "gbpusd", "usdchf", "usdjpy", "usdcad", "audusd", "nzdusd",
            "eurjpy", "gbpjpy", "audjpy", "cadjpy", "chfjpy", "nzdjpy",
            "eurgbp", "euraud", "eurcad", "eurchf", "eurnzd",
            "gbpaud", "gbpcad", "gbpchf", "gbpnzd",
            "audcad", "audchf", "audnzd", "cadchf", "nzdcad", "nzdchf"
        ]
        
        self.symboles_metaux = ["gold", "xauusd", "xau", "or", "silver", "xagusd", "xag", "argent", "platinum", "xptusd", "palladium", "xpdusd"]
        self.symboles_indices = ["dax", "cac", "sp500", "dow", "nasdaq", "ftse", "nikkei", "asx", "us30", "us500", "ger30", "fra40", "uk100", "ger40"]
        self.symboles_crypto = ["btc", "eth", "ltc", "xrp", "ada", "dot", "bitcoin", "ethereum", "crypto"]
        self.symboles_energie = ["oil", "wti", "brent", "petrol", "crude", "gas", "natural"]
    
    def process_files(self, file_paths, task_id, task_status, filter_type=None):
        """
        Traite une liste de fichiers Excel
        filter_type: 'forex', 'autres', ou None (tous)
        """
        try:
            print(f"[DEBUG] Starting unified analysis with {len(file_paths)} files, filter: {filter_type}")
            tous_les_resultats = []
            total_files = len(file_paths)
            
            for i, file_path in enumerate(file_paths):
                progress = 20 + (i / total_files) * 40
                task_status[task_id]['progress'] = int(progress)
                task_status[task_id]['message'] = f'Traitement du fichier {i+1}/{total_files}...'
                
                print(f"[DEBUG] Processing file {i+1}/{total_files}: {os.path.basename(file_path)}")
                
                df_result, erreur, exclus, doublons = self.process_single_file(file_path, filter_type)
                
                if df_result is not None and len(df_result) > 0:
                    tous_les_resultats.append(df_result)
                    
                    filename = os.path.basename(file_path)
                    # Compter les trades complets (clés uniques) au lieu des opérations
                    nb_trades_complets = df_result["Cle_Match"].nunique() if "Cle_Match" in df_result.columns else len(df_result)
                    self.statistiques_fichiers[filename] = {
                        'trades': nb_trades_complets,
                        'exclus': exclus,
                        'doublons': doublons,
                        'erreur': erreur
                    }
                    print(f"[DEBUG] File processed successfully: {len(df_result)} trades, {exclus} excluded")
                else:
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': 0,
                        'exclus': 0,
                        'doublons': 0,
                        'erreur': erreur or "Aucune donnée trouvée"
                    }
                    print(f"[DEBUG] File failed: {erreur}")
            
            if not tous_les_resultats:
                print(f"[DEBUG] No valid data found in any file")
                return None
            
            task_status[task_id]['progress'] = 60
            task_status[task_id]['message'] = 'Fusion des données et calculs des intérêts composés...'
            
            print(f"[DEBUG] Starting fusion and compound interest calculations")
            df_final = self.fusionner_et_calculer_cumuls(tous_les_resultats)
            print(f"[DEBUG] Fusion completed: {len(df_final)} total trades")
            
            task_status[task_id]['progress'] = 75
            task_status[task_id]['message'] = 'Calculs des statistiques avancées...'
            
            return df_final
            
        except Exception as e:
            print(f"[ERROR] Error in process_files: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors du traitement des fichiers: {str(e)}")
    
    def process_single_file(self, file_path, filter_type=None):
        """Traite un seul fichier Excel avec filtrage optionnel"""
        try:
            print(f"[DEBUG] Starting to process file: {file_path}")
            df = pd.read_excel(file_path, sheet_name=0, header=None)
            print(f"[DEBUG] File read successfully, shape: {df.shape}")
            
            # Trouver les lignes "ordre" et "transaction"
            ligne_ordres = self.trouver_ligne(df, "ordre")
            ligne_transactions = self.trouver_ligne(df, "transaction")
            print(f"[DEBUG] Found ordre line at: {ligne_ordres}, transaction line at: {ligne_transactions}")

            # Extraire les DataFrames ordres et transactions
            header_ordres = df.iloc[ligne_ordres + 1]
            ordres_df = df.iloc[ligne_ordres + 2 : ligne_transactions].copy()
            ordres_df.columns = header_ordres
            ordres_df.reset_index(drop=True, inplace=True)

            header_transactions = df.iloc[ligne_transactions + 1]
            transactions_df = df.iloc[ligne_transactions + 2 :].copy()
            transactions_df.columns = header_transactions
            transactions_df = transactions_df[transactions_df.iloc[:, 0].notna()]
            transactions_df.reset_index(drop=True, inplace=True)
            
            print(f"[DEBUG] Ordres shape: {ordres_df.shape}, Transactions shape: {transactions_df.shape}")

            if len(ordres_df.columns) < 2 or len(transactions_df.columns) < 2:
                return None, "Pas assez de colonnes dans les données", 0, 0

            # Créer la clé de jointure
            ordres_df["__clé__"] = ordres_df.iloc[:, 1].astype(str)
            transactions_df["__clé__"] = transactions_df.iloc[:, 1].astype(str)

            # Ajouter la colonne d'origine du fichier pour désambiguïser les clés
            fichier_source = os.path.basename(file_path)
            ordres_df["Fichier_Source"] = fichier_source
            transactions_df["Fichier_Source"] = fichier_source
            
            # Renommer la colonne Prix si elle existe
            if "Prix" in transactions_df.columns:
                transactions_df.rename(columns={"Prix": "Prix_transaction"}, inplace=True)

            # Fusionner les DataFrames
            fusion_df = pd.merge(ordres_df, transactions_df, on="__clé__", suffixes=('_ordre', '_transaction'))
            print(f"[DEBUG] Merged dataframe shape: {fusion_df.shape}")

            # Unifier la colonne Fichier_Source après merge
            if "Fichier_Source_ordre" in fusion_df.columns:
                fusion_df["Fichier_Source"] = fusion_df["Fichier_Source_ordre"]
            elif "Fichier_Source_transaction" in fusion_df.columns:
                fusion_df["Fichier_Source"] = fusion_df["Fichier_Source_transaction"]
            # Nettoyer les colonnes intermédiaires si présentes
            colonnes_a_supprimer_tmp = []
            for col_tmp in ["Fichier_Source_ordre", "Fichier_Source_transaction"]:
                if col_tmp in fusion_df.columns:
                    colonnes_a_supprimer_tmp.append(col_tmp)
            if colonnes_a_supprimer_tmp:
                fusion_df.drop(columns=colonnes_a_supprimer_tmp, inplace=True)
            
            avant_filtrage = len(fusion_df)

            # Filtrage selon le type demandé
            apres_filtrage = avant_filtrage  # Initialisation par défaut
            
            if "Symbole_ordre" in fusion_df.columns and filter_type:
                print(f"[DEBUG] Applying {filter_type} filter...")
                if filter_type == 'forex':
                    fusion_df = fusion_df[fusion_df["Symbole_ordre"].apply(self.est_forex)]
                elif filter_type == 'autres':
                    fusion_df = fusion_df[fusion_df["Symbole_ordre"].apply(self.est_autre_instrument)]
                
                apres_filtrage = len(fusion_df)
                print(f"[DEBUG] After filtering: {apres_filtrage} rows (excluded: {avant_filtrage - apres_filtrage})")
                
                if len(fusion_df) == 0:
                    return None, f"Aucun instrument {filter_type} trouvé", avant_filtrage - apres_filtrage, 0

            # Conversions des colonnes numériques
            print(f"[DEBUG] Converting numeric columns...")
            fusion_df["Profit"] = self.safe_convert_to_float(fusion_df["Profit"])
            fusion_df["Prix_transaction"] = self.safe_convert_to_float(fusion_df["Prix_transaction"])
            
            if "T / P" in fusion_df.columns:
                fusion_df["T / P"] = self.safe_convert_to_float(fusion_df["T / P"])
            if "S / L" in fusion_df.columns:
                fusion_df["S / L"] = self.safe_convert_to_float(fusion_df["S / L"])

            fusion_df["Volume_ordre"] = fusion_df["Volume_ordre"].astype(str)
            fusion_df["Symbole_ordre"] = fusion_df["Symbole_ordre"].astype(str)
            fusion_df["Cle_Match"] = None

            # Logique de matching des trades
            print(f"[DEBUG] Applying matching logic...")
            self.apply_matching_logic(fusion_df)
            
            # Créer l'index des trades d'entrée
            df_in = fusion_df[(fusion_df["Direction"] == "in") & (fusion_df["Cle_Match"].notna())].copy()
            if len(df_in) > 0:
                df_in = df_in.set_index("Cle_Match")

            # Calcul des pips/points selon le type d'instrument
            print(f"[DEBUG] Calculating pips/points...")
            fusion_df["Profit_pips"] = fusion_df.apply(lambda row: self.calculer_pips_ou_points(row, df_in), axis=1)
            
            # Nettoyage et sélection des colonnes finales
            colonnes_a_garder = [
                "Heure d'ouverture", "Ordre_ordre", "Symbole_ordre", "Type_ordre", 
                "Volume_ordre", "S / L", "T / P", "Direction", "Prix_transaction",
                "Profit", "Cle_Match", "Profit_pips", "Fichier_Source"
            ]
            
            colonnes_finales = [col for col in colonnes_a_garder if col in fusion_df.columns]
            fusion_df = fusion_df[colonnes_finales]
            
            # Suppression des doublons
            avant_dedoublonnage = len(fusion_df)
            fusion_df = fusion_df.drop_duplicates().reset_index(drop=True)
            apres_dedoublonnage = len(fusion_df)
            doublons_supprimes = avant_dedoublonnage - apres_dedoublonnage
            
            print(f"[DEBUG] File processing completed: {len(fusion_df)} final trades")
            
            # Calculer le nombre d'exclus
            exclus = avant_filtrage - apres_filtrage
            
            return fusion_df, "Succès", exclus, doublons_supprimes
            
        except Exception as e:
            print(f"[ERROR] Error processing file {file_path}: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            return None, str(e), 0, 0
    
    def trouver_ligne(self, df, mot_approx):
        """Trouve une ligne contenant un mot approximatif"""
        for i, row in df.iterrows():
            texte = row.astype(str).str.lower().str.replace(" ", "").str.replace(":", "")
            if texte.str.contains(mot_approx.lower()).any():
                return i
        raise ValueError(f"Ligne avec '{mot_approx}' non trouvée.")
    
    def safe_convert_to_float(self, series):
        """Convertit une série en float en gérant les valeurs NaN"""
        return pd.to_numeric(series.astype(str).str.replace(",", ".").replace("nan", ""), errors='coerce')
    
    def detecter_type_instrument(self, symbole):
        """Détecte automatiquement le type d'instrument financier"""
        symbole = str(symbole).upper()
        
        # Reconnaissance automatique Forex
        if self.est_forex_automatique(symbole):
            return InstrumentType.FOREX
        
        # Reconnaissance automatique Métaux
        if self.est_metal_automatique(symbole):
            return InstrumentType.METAUX
        
        # Reconnaissance automatique Indices
        if self.est_indice_automatique(symbole):
            return InstrumentType.INDICES
        
        # Reconnaissance automatique Crypto
        if self.est_crypto_automatique(symbole):
            return InstrumentType.CRYPTO
        
        # Reconnaissance automatique Énergie
        if self.est_energie_automatique(symbole):
            return InstrumentType.ENERGIE
        
        # Fallback sur les listes existantes
        symbole_lower = symbole.lower()
        if any(metal in symbole_lower for metal in self.symboles_metaux):
            return InstrumentType.METAUX
        elif any(index in symbole_lower for index in self.symboles_indices):
            return InstrumentType.INDICES
        elif any(c in symbole_lower for c in self.symboles_crypto):
            return InstrumentType.CRYPTO
        elif any(e in symbole_lower for e in self.symboles_energie):
            return InstrumentType.ENERGIE
        elif any(forex_pair in symbole_lower for forex_pair in self.symboles_forex):
            return InstrumentType.FOREX
        else:
            return InstrumentType.ACTIONS
    
    def est_forex_automatique(self, symbole):
        """Reconnaissance automatique des paires Forex"""
        # Devises connues
        devises = ["USD", "EUR", "GBP", "JPY", "CHF", "CAD", "AUD", "NZD"]
        
        # Pattern Forex : 6 lettres (3+3) avec des devises connues
        if len(symbole) == 6:
            devise1 = symbole[:3]
            devise2 = symbole[3:]
            if devise1 in devises and devise2 in devises:
                return True
        
        # Pattern Forex : 7 lettres (3+4) comme EURJPY
        if len(symbole) == 7:
            devise1 = symbole[:3]
            devise2 = symbole[3:]
            if devise1 in devises and devise2 in devises:
                return True
        
        return False
    
    def est_metal_automatique(self, symbole):
        """Reconnaissance automatique des métaux"""
        mots_cles_metaux = ["GOLD", "SILVER", "XAU", "XAG", "PLATINUM", "PALLADIUM"]
        return any(mot in symbole for mot in mots_cles_metaux)
    
    def est_indice_automatique(self, symbole):
        """Reconnaissance automatique des indices"""
        mots_cles_indices = ["DAX", "CAC", "SP500", "NASDAQ", "FTSE", "NIKKEI", "DOW"]
        patterns_indices = ["US30", "US500", "GER30", "FRA40", "UK100", "GER40"]
        
        return any(mot in symbole for mot in mots_cles_indices) or any(pattern in symbole for pattern in patterns_indices)
    
    def est_crypto_automatique(self, symbole):
        """Reconnaissance automatique des cryptos"""
        mots_cles_crypto = ["BTC", "ETH", "LTC", "XRP", "ADA", "DOT", "BITCOIN", "ETHEREUM"]
        return any(mot in symbole for mot in mots_cles_crypto)
    
    def est_energie_automatique(self, symbole):
        """Reconnaissance automatique des énergies"""
        mots_cles_energie = ["OIL", "WTI", "BRENT", "GAS", "NATURAL"]
        return any(mot in symbole for mot in mots_cles_energie)
    
    def est_forex(self, symbole):
        """Vérifie si un symbole est une paire Forex"""
        symbole = str(symbole).lower()
        return any(forex_pair in symbole for forex_pair in self.symboles_forex)
    
    def est_autre_instrument(self, symbole):
        """Vérifie si un symbole N'EST PAS une paire Forex"""
        return not self.est_forex(symbole)
    
    def apply_matching_logic(self, fusion_df):
        """LOGIQUE DE MATCHING RESPECTANT TOUS LES CRITÈRES - N° ordre, volumes, TP/SL"""
        print(f"[DEBUG] Starting CRITERIA-BASED matching logic for {len(fusion_df)} rows")
        
        # TRIER CHRONOLOGIQUEMENT AVANT LE MATCHING
        if "Date_ordre" in fusion_df.columns:
            fusion_df = fusion_df.sort_values("Date_ordre").reset_index(drop=True)
            print(f"[DEBUG] DataFrame trié chronologiquement par Date_ordre")
        elif "Date_transaction" in fusion_df.columns:
            fusion_df = fusion_df.sort_values("Date_transaction").reset_index(drop=True)
            print(f"[DEBUG] DataFrame trié chronologiquement par Date_transaction")
        else:
            print(f"[WARNING] Aucune colonne de date trouvée pour le tri chronologique")
        
        # ÉTAPE 1: Créer les clés pour TOUS les trades "in"
        trades_in = fusion_df[fusion_df["Direction"] == "in"].copy()
        print(f"[DEBUG] Creating keys for {len(trades_in)} 'in' trades")
        
        for idx, row in trades_in.iterrows():
            ordre = row["Ordre_ordre"]
            symbole = row["Symbole_ordre"]
            fichier = row.get("Fichier_Source", "")
            
            # Créer une clé UNIQUE basée sur le fichier + symbole + ordre
            cle_unique = f"{fichier}|{symbole}-{ordre}"
            fusion_df.at[idx, "Cle_Match"] = cle_unique
            print(f"[DEBUG] Created key for {fichier} / {symbole} order {ordre}: {cle_unique}")

        # ÉTAPE 2: MATCHING RESPECTANT TOUS LES CRITÈRES
        print(f"[DEBUG] Starting CRITERIA-BASED matching for all symbols")
        
        # Grouper par (symbole, fichier) pour éviter toute collision inter-fichiers
        if "Fichier_Source" in fusion_df.columns:
            groupes = fusion_df[["Symbole_ordre", "Fichier_Source"]].drop_duplicates()
            iter_groupes = [(row.Symbole_ordre, row.Fichier_Source) for _, row in groupes.iterrows()]
        else:
            iter_groupes = [(symb, None) for symb in fusion_df["Symbole_ordre"].unique()]

        for symbole, fichier in iter_groupes:
            print(f"[DEBUG] Processing symbol: {symbole} (file: {fichier})")
            
            # Récupérer tous les trades pour ce symbole
            if fichier is not None:
                trades_symbole = fusion_df[(fusion_df["Symbole_ordre"] == symbole) & (fusion_df["Fichier_Source"] == fichier)].copy()
            else:
                trades_symbole = fusion_df[fusion_df["Symbole_ordre"] == symbole].copy()
            trades_in_symbole = trades_symbole[trades_symbole["Direction"] == "in"].copy()
            trades_out_symbole = trades_symbole[trades_symbole["Direction"] == "out"].copy()
            
            print(f"[DEBUG] {symbole}: {len(trades_in_symbole)} IN trades, {len(trades_out_symbole)} OUT trades")
            
            # Pour chaque trade "in", chercher TOUS les trades "out" correspondants
            for _, trade_in in trades_in_symbole.iterrows():
                ordre_in = trade_in["Ordre_ordre"]
                volume_in = self.parse_volume(trade_in["Volume_ordre"])
                cle_in = trade_in["Cle_Match"]
                tp_in = trade_in.get("T / P", None)
                sl_in = trade_in.get("S / L", None)
                
                print(f"[DEBUG] Processing IN trade: {symbole} order {ordre_in}, volume: {volume_in}")
                
                # Nouvelle logique: agréger séquentiellement les OUT jusqu'à atteindre le volume IN
                candidates_out = trades_out_symbole[
                    (trades_out_symbole["Ordre_ordre"] > ordre_in) &
                    (trades_out_symbole["Cle_Match"].isna())
                ].copy()

                if len(candidates_out) == 0:
                    print(f"[DEBUG] No OUT trades after IN {ordre_in}")
                    continue

                candidates_out = candidates_out.sort_values("Ordre_ordre").reset_index(drop=True)

                cum_volume = 0.0
                selected_out_rows = []
                for _, trade_out in candidates_out.iterrows():
                    volume_out = self.parse_volume(trade_out["Volume_ordre"])
                    cum_volume += volume_out
                    selected_out_rows.append(trade_out)
                    if math.isclose(cum_volume, volume_in, rel_tol=0.02, abs_tol=1e-6) or cum_volume > volume_in:
                        break

                print(f"[DEBUG] Aggregated OUT volume: {cum_volume} for IN {volume_in} with {len(selected_out_rows)} outs")

                if math.isclose(cum_volume, volume_in, rel_tol=0.02, abs_tol=1e-6):
                    print(f"[DEBUG] ✅ Aggregated volume matches IN within tolerance. Assigning {len(selected_out_rows)} OUT trades")
                    for trade_out in selected_out_rows:
                        if fichier is not None:
                            idx_out = fusion_df[
                                (fusion_df["Symbole_ordre"] == symbole) &
                                (fusion_df["Ordre_ordre"] == trade_out["Ordre_ordre"]) &
                                (fusion_df["Direction"] == "out") &
                                (fusion_df["Fichier_Source"] == fichier)
                            ].index[0]
                        else:
                            idx_out = fusion_df[
                                (fusion_df["Symbole_ordre"] == symbole) &
                                (fusion_df["Ordre_ordre"] == trade_out["Ordre_ordre"]) &
                                (fusion_df["Direction"] == "out")
                            ].index[0]
                        fusion_df.at[idx_out, "Cle_Match"] = cle_in
                        print(f"[DEBUG] ✅ Assigned {cle_in} to {symbole} order {trade_out['Ordre_ordre']}")
                else:
                    print(f"[DEBUG] ❌ Aggregated volume does not match. Trying 1:1 fallback")
                    # Fallback: 1:1 sur le premier OUT proche
                    for _, trade_out in candidates_out.iterrows():
                        volume_out = self.parse_volume(trade_out["Volume_ordre"])
                        if math.isclose(volume_out, volume_in, rel_tol=0.02, abs_tol=1e-6):
                            print(f"[DEBUG] 🔄 Fallback: Simple 1:1 match found")
                            if fichier is not None:
                                idx_out = fusion_df[
                                    (fusion_df["Symbole_ordre"] == symbole) &
                                    (fusion_df["Ordre_ordre"] == trade_out["Ordre_ordre"]) &
                                    (fusion_df["Direction"] == "out") &
                                    (fusion_df["Fichier_Source"] == fichier)
                                ].index[0]
                            else:
                                idx_out = fusion_df[
                                    (fusion_df["Symbole_ordre"] == symbole) &
                                    (fusion_df["Ordre_ordre"] == trade_out["Ordre_ordre"]) &
                                    (fusion_df["Direction"] == "out")
                                ].index[0]
                            fusion_df.at[idx_out, "Cle_Match"] = cle_in
                            break
        
        # ÉTAPE 3: VÉRIFICATION FINALE - Assigner les trades OUT restants
        print(f"[DEBUG] Final verification and cleanup")
        
        trades_out_sans_cle = fusion_df[(fusion_df["Direction"] == "out") & (fusion_df["Cle_Match"].isna())]
        if len(trades_out_sans_cle) > 0:
            print(f"[DEBUG] ⚠️ {len(trades_out_sans_cle)} OUT trades still without key - applying emergency matching")
            
            for _, trade_out in trades_out_sans_cle.iterrows():
                symbole = trade_out["Symbole_ordre"]
                ordre_out = trade_out["Ordre_ordre"]
                fichier = trade_out.get("Fichier_Source", None)
                
                # Chercher le trade IN le plus proche (avant ou après)
                if fichier is not None:
                    trades_in_symbole = fusion_df[
                        (fusion_df["Symbole_ordre"] == symbole) & 
                        (fusion_df["Direction"] == "in") &
                        (fusion_df["Fichier_Source"] == fichier)
                    ].copy()
                else:
                    trades_in_symbole = fusion_df[
                        (fusion_df["Symbole_ordre"] == symbole) & 
                        (fusion_df["Direction"] == "in")
                    ].copy()
                
                if len(trades_in_symbole) > 0:
                    # Trouver le trade IN le plus proche en numéro d'ordre
                    trades_in_symbole["Distance"] = abs(trades_in_symbole["Ordre_ordre"] - ordre_out)
                    trade_in_proche = trades_in_symbole.loc[trades_in_symbole["Distance"].idxmin()]
                    
                    print(f"[DEBUG] 🚨 Emergency match: OUT {ordre_out} -> IN {trade_in_proche['Ordre_ordre']}")
                    
                    if fichier is not None:
                        idx_out = fusion_df[
                            (fusion_df["Symbole_ordre"] == symbole) & 
                            (fusion_df["Ordre_ordre"] == ordre_out) & 
                            (fusion_df["Direction"] == "out") &
                            (fusion_df["Fichier_Source"] == fichier)
                        ].index[0]
                    else:
                        idx_out = fusion_df[
                            (fusion_df["Symbole_ordre"] == symbole) & 
                            (fusion_df["Ordre_ordre"] == ordre_out) & 
                            (fusion_df["Direction"] == "out")
                        ].index[0]
                    fusion_df.at[idx_out, "Cle_Match"] = trade_in_proche["Cle_Match"]
        
        # RÉSULTATS FINAUX
        trades_in_avec_cle = len(fusion_df[(fusion_df["Direction"] == "in") & (fusion_df["Cle_Match"].notna())])
        trades_out_avec_cle = len(fusion_df[(fusion_df["Direction"] == "out") & (fusion_df["Cle_Match"].notna())])
        total_avec_cle = trades_in_avec_cle + trades_out_avec_cle
        
        print(f"[DEBUG] 🎯 CRITERIA-BASED MATCHING RESULTS:")
        print(f"[DEBUG]   - IN trades with keys: {trades_in_avec_cle}")
        print(f"[DEBUG]   - OUT trades with keys: {trades_out_avec_cle}")
        print(f"[DEBUG]   - Total with keys: {total_avec_cle}/{len(fusion_df)} ({(total_avec_cle/len(fusion_df)*100):.1f}%)")
        
        # Vérification finale
        if total_avec_cle == len(fusion_df):
            print(f"[DEBUG] ✅ SUCCESS: ALL trades have keys!")
        else:
            print(f"[DEBUG] ❌ WARNING: {len(fusion_df) - total_avec_cle} trades still without keys!")
    
    def parse_volume(self, volume_str):
        """Parse le volume depuis une chaîne comme '0.56 / 0.56' ou '0.56'"""
        try:
            volume_str = str(volume_str).strip()
            if "/" in volume_str:
                # Format "executed / total"
                return float(volume_str.split("/")[0].strip())
            else:
                return float(volume_str)
        except:
            return 0.0
    
    def extract_price_from_comment(self, commentaire):
        """Extrait le prix d'un commentaire TP/SL"""
        import re
        commentaire = str(commentaire).lower()
        match = re.search(r'(tp|sl)[^\d]*(\d+[.,]?\d*)', commentaire)
        if match:
            try:
                return float(match.group(2).replace(",", "."))
            except:
                return 0.0
        return 0.0
    
    def extraire_prix_commentaire(self, commentaire):
        """Extrait le prix du commentaire"""
        commentaire = str(commentaire).lower()
        match = re.search(r'(tp|sl)[^\d]*(\d+[.,]?\d+)', commentaire)
        if match:
            try:
                prix = float(match.group(2).replace(",", "."))
                return round(prix, 5)
            except:
                return None
        return None
    
    def calculer_pips_ou_points(self, row, df_in):
        """Calcul des pips (Forex) ou points (autres instruments) avec détection d'incohérences"""
        symbole = str(row["Symbole_ordre"]).lower()
        profit = row["Profit"]
        type_instrument = self.detecter_type_instrument(symbole)
        
        # Gestion du volume
        volume_str = str(row["Volume_ordre"])
        if "/" in volume_str:
            volume = float(volume_str.split("/")[0].strip())
        else:
            volume = float(volume_str.strip())
        
        try:
            # Si c'est un trade de sortie avec matching
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points_bruts = prix_out - prix_in
                        else:
                            points_bruts = prix_in - prix_out
                        
                        # Conversion pips/points selon règles demandées
                        if type_instrument == InstrumentType.FOREX:
                            # Déterminer la taille de pip à partir du nombre de décimales des prix
                            def _decimals(x):
                                try:
                                    s = f"{float(x):.10f}".rstrip("0").rstrip(".")
                                    return len(s.split(".")[1]) if "." in s else 0
                                except Exception:
                                    return 0
                            nb_dec = max(_decimals(prix_in), _decimals(prix_out))
                            if nb_dec >= 4:
                                pip_size = 0.0001  # 4e décimale, 5e ignorée (pipette)
                            elif nb_dec in (2, 3):
                                pip_size = 0.01    # 2e décimale
                            else:
                                # Fallback: par défaut considérer 0.0001
                                pip_size = 0.0001

                            pips_floats = abs(points_bruts) / pip_size
                            pips_entiers = int(pips_floats)  # ignorer les pipettes (pipettes non comptées)
                            signe = 1 if points_bruts >= 0 else -1
                            return signe * pips_entiers
                        else:
                            # Pour les autres instruments, conserver les points bruts
                            return round(points_bruts, 2)
            
            # Fallback : calcul basé sur le profit avec valeurs réalistes
            if type_instrument == InstrumentType.FOREX:
                # Valeurs réalistes pour le Forex (basées sur des spreads typiques)
                # On garde l'approximation de 10€/pip par 0.1 lot comme fallback
                valeur_pip = volume * 10.0
                
                if valeur_pip != 0:
                    pips_calcules = round(profit / valeur_pip, 2)
                    print(f"[INFO] Calcul fallback pour {symbole}: {profit}€ / {valeur_pip}€ = {pips_calcules} pips")
                    return pips_calcules
            else:
                # Valeurs réalistes pour les autres instruments
                if type_instrument == InstrumentType.METAUX:
                    # Or/Argent : 1 point = ~1€ pour 0.1 lot
                    valeur_point = volume * 1.0
                elif type_instrument == InstrumentType.INDICES:
                    if "dax" in symbole or "ger30" in symbole or "ger40" in symbole:
                        # DAX/GER30/GER40 : 1 point = ~1€ pour 0.1 lot
                        valeur_point = volume * 1.0
                    elif "cac" in symbole or "fra40" in symbole:
                        # CAC40 : 1 point = ~1€ pour 0.1 lot
                        valeur_point = volume * 1.0
                    elif "sp500" in symbole or "us500" in symbole:
                        # SP500 : 1 point = ~1€ pour 0.1 lot
                        valeur_point = volume * 1.0
                    else:
                        # Autres indices : 1 point = ~1€ pour 0.1 lot
                        valeur_point = volume * 1.0
                elif type_instrument == InstrumentType.CRYPTO:
                    # Crypto : 1 point = ~0.1€ pour 0.1 lot
                    valeur_point = volume * 0.1
                elif type_instrument == InstrumentType.ENERGIE:
                    # Pétrole : 1 point = ~1€ pour 0.1 lot
                    valeur_point = volume * 1.0
                else:
                    # Actions : 1 point = ~1€ pour 0.1 lot
                    valeur_point = volume * 1.0
                
                if valeur_point != 0:
                    return round(profit / valeur_point, 2)
            
            return None
                
        except Exception as e:
            print(f"[ERROR] Erreur calcul pips pour {symbole}: {str(e)}")
            return None
    
    def fusionner_et_calculer_cumuls(self, tous_les_df):
        """Fusionne tous les DataFrames et calcule les intérêts composés + drawdown"""
        print(f"[DEBUG] Starting fusion and compound calculations...")
        
        # Fusionner tous les DataFrames
        df_complet = pd.concat(tous_les_df, ignore_index=True)
        print(f"[DEBUG] Merged {len(tous_les_df)} dataframes into {len(df_complet)} total trades")
        
        # Tri par date
        if "Heure d'ouverture" in df_complet.columns:
            df_complet["Date_parsed"] = pd.to_datetime(df_complet["Heure d'ouverture"], errors='coerce')
            df_complet = df_complet.sort_values("Date_parsed").reset_index(drop=True)
            df_complet = df_complet.drop("Date_parsed", axis=1)
        
        # Calculs cumulés avec intérêts composés
        df_complet["Profit_compose"] = 0.0
        df_complet["Profit_cumule"] = 0.0
        df_complet["Solde_cumule"] = 0.0
        df_complet["Profit_pips_cumule"] = 0.0
        df_complet["Drawdown_pct"] = 0.0
        df_complet["Drawdown_euros"] = 0.0
        df_complet["Drawdown_running_pct"] = 0.0
        
        solde_courant = self.solde_initial
        profit_cumule_reel = 0.0
        pips_cumule = 0.0
        plus_haut_solde = self.solde_initial
        drawdown_running_max = 0.0
        
        print(f"[DEBUG] Starting compound interest calculations...")
        
        for idx, row in df_complet.iterrows():
            profit_original = row["Profit"] if pd.notna(row["Profit"]) else 0
            pips = row["Profit_pips"] if pd.notna(row["Profit_pips"]) else 0
            
            # Calculer le rendement en pourcentage
            if profit_original != 0 and self.solde_initial != 0:
                rendement_trade_pct = (profit_original / self.solde_initial) * 100
                profit_compose = (rendement_trade_pct / 100) * solde_courant
            else:
                profit_compose = 0
            
            # Mise à jour des cumuls
            solde_courant += profit_compose
            profit_cumule_reel += profit_compose
            pips_cumule += pips
            
            # Mise à jour du plus haut solde historique
            if solde_courant > plus_haut_solde:
                plus_haut_solde = solde_courant
            
            # Calcul du drawdown CLASSIQUE
            if solde_courant < plus_haut_solde:
                drawdown_euros = plus_haut_solde - solde_courant
                drawdown_pct = (drawdown_euros / plus_haut_solde * 100)
            else:
                drawdown_euros = 0.0
                drawdown_pct = 0.0
            
            # Calcul du drawdown RUNNING (lissé)
            drawdown_actuel = (plus_haut_solde - solde_courant) / plus_haut_solde * 100
            if drawdown_actuel > drawdown_running_max:
                drawdown_running_max = drawdown_actuel
            
            if drawdown_actuel < drawdown_running_max:
                if profit_original > 0:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max * 0.9)
                else:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max)
            
            # Enregistrer les valeurs
            df_complet.at[idx, "Profit_compose"] = round(profit_compose, 2)
            df_complet.at[idx, "Profit_cumule"] = round(profit_cumule_reel, 2)
            df_complet.at[idx, "Solde_cumule"] = round(solde_courant, 2)
            df_complet.at[idx, "Profit_pips_cumule"] = round(pips_cumule, 2)
            df_complet.at[idx, "Drawdown_pct"] = round(drawdown_pct, 2)
            df_complet.at[idx, "Drawdown_euros"] = round(drawdown_euros, 2)
            df_complet.at[idx, "Drawdown_running_pct"] = round(drawdown_running_max, 2)
        
        print(f"[DEBUG] Compound calculations completed. Final solde: {solde_courant:.2f}")
        print(f"[DEBUG] Max drawdown: {df_complet['Drawdown_pct'].max():.2f}%")
        
        return df_complet
    
    def calculer_trades_complets(self, df):
        """Calcule le nombre de trades complets (1 IN + 1 ou plusieurs OUT)"""
        print(f"[DEBUG] Calculating complete trades from {len(df)} rows")
        
        # Compter les clés de jointure uniques (chaque clé = 1 trade complet)
        trades_complets = df["Cle_Match"].nunique()
        
        # Compter les trades IN et OUT uniques (pour debug)
        trades_in_uniques = df[df["Direction"] == "in"]["Cle_Match"].nunique()
        trades_out_uniques = df[df["Direction"] == "out"]["Cle_Match"].nunique()
        
        print(f"[DEBUG] Found {trades_complets} complete trades (unique keys), {trades_in_uniques} unique IN trades, {trades_out_uniques} unique OUT trades")
        
        # Le nombre de trades complets = nombre de clés de jointure uniques
        return trades_complets
    
    def calculer_trades_par_resultat(self, df):
        """Calcule les trades gagnants/perdants basés sur les trades complets"""
        print(f"[DEBUG] Calculating trades by result")
        
        # Grouper par clé de trade et calculer le profit total de chaque trade complet
        trades_complets = df.groupby("Cle_Match").agg({
            "Profit": "sum"
        }).reset_index()
        
        # Compter par résultat (tous les trades complets, pas seulement les IN)
        trades_gagnants = len(trades_complets[trades_complets["Profit"] > 0])
        trades_perdants = len(trades_complets[trades_complets["Profit"] < 0])
        trades_neutres = len(trades_complets[trades_complets["Profit"] == 0])
        
        print(f"[DEBUG] Complete trades: {len(trades_complets)} total, {trades_gagnants} winners, {trades_perdants} losers")
        
        return trades_gagnants, trades_perdants, trades_neutres, len(trades_complets)
    
    def calculer_statistiques_avancees(self, df):
        """Calcule les statistiques avancées basées sur les trades complets"""
        stats = {}
        
        # Calculer les trades complets par résultat
        trades_gagnants, trades_perdants, trades_neutres, total_trades_complets = self.calculer_trades_par_resultat(df)
        
        # Calculer les profits moyens basés sur les trades complets
        trades_complets = df.groupby("Cle_Match").agg({
            "Profit": "sum"
        }).reset_index()
        
        profits_gagnants = trades_complets[trades_complets["Profit"] > 0]["Profit"]
        profits_perdants = trades_complets[trades_complets["Profit"] < 0]["Profit"]
        
        # Moyennes basées sur les trades complets
        stats["gain_moyen"] = profits_gagnants.mean() if len(profits_gagnants) > 0 else 0
        stats["perte_moyenne"] = profits_perdants.mean() if len(profits_perdants) > 0 else 0
        
        # Calcul des séries consécutives basées sur les trades complets
        series_gagnantes = []
        series_perdantes = []
        
        serie_gagnante_actuelle = 0
        serie_perdante_actuelle = 0
        
        # Trier par ordre chronologique pour les séries
        if "Heure d'ouverture" in df.columns:
            df_triee = df.sort_values("Heure d'ouverture")
        elif "Ordre_ordre" in df.columns:
            df_triee = df.sort_values("Ordre_ordre")
        else:
            df_triee = df
        
        # Parcourir les trades complets dans l'ordre chronologique
        for _, row in trades_complets.iterrows():
            profit = row["Profit"]
            
            if profit > 0:
                serie_gagnante_actuelle += 1
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
            elif profit < 0:
                serie_perdante_actuelle += 1
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
            else:
                # Trade neutre
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
        
        # Ajouter la dernière série
        if serie_gagnante_actuelle > 0:
            series_gagnantes.append(serie_gagnante_actuelle)
        if serie_perdante_actuelle > 0:
            series_perdantes.append(serie_perdante_actuelle)
        
        stats["gains_consecutifs_max"] = max(series_gagnantes) if series_gagnantes else 0
        stats["pertes_consecutives_max"] = max(series_perdantes) if series_perdantes else 0
        
        # Statistiques du drawdown (basées sur toutes les lignes pour la continuité)
        stats["drawdown_max_pct"] = df["Drawdown_pct"].max()
        stats["drawdown_max_euros"] = df["Drawdown_euros"].max()
        
        # Nombre de périodes de drawdown
        periodes_drawdown = len(df[df["Drawdown_pct"] > 0])
        stats["periodes_drawdown"] = periodes_drawdown
        
        # Ajouter les statistiques de trades complets
        stats["total_trades_complets"] = total_trades_complets
        stats["trades_gagnants_complets"] = trades_gagnants
        stats["trades_perdants_complets"] = trades_perdants
        stats["trades_neutres_complets"] = trades_neutres
        
        return stats

    def calculer_agregations_graphes(self, df: pd.DataFrame):
        """Calcule les agrégations nécessaires pour les graphiques demandés.

        - Horaires d'ouverture majoritaires (IN)
        - Horaires de fermeture majoritaires (dernier OUT par trade)
        - Heures/Jours/Mois: profits totaux survenus (sur les OUT)
        - Comptes TP/SL par heure/jour/mois (TP=profit total du trade >0, SL<0, au dernier OUT)
        - Évolution cumulée du profit (linéaire) au moment du dernier OUT de chaque trade
        - Durée moyenne/médiane des trades (IN -> dernier OUT)
        """
        result = {}

        if "Heure d'ouverture" not in df.columns or "Direction" not in df.columns:
            return result

        # Normaliser la date
        df = df.copy()
        df["Datetime"] = pd.to_datetime(df["Heure d'ouverture"], errors='coerce')
        df = df[df["Datetime"].notna()]

        # Préparer tables IN et OUT
        df_in = df[df["Direction"] == "in"].copy()
        df_out = df[df["Direction"] == "out"].copy()

        # 1) Horaires d'ouvertures (IN)
        heures_index = pd.Index(range(24), name="Heure")
        if len(df_in) > 0:
            heures_in = df_in["Datetime"].dt.hour.value_counts().sort_index()
            heures_in = heures_in.reindex(heures_index, fill_value=0)
        else:
            heures_in = pd.Series(0, index=heures_index, dtype=int)
        result["heures_in_counts"] = heures_in

        # 2) Dernier OUT par trade (fermeture)
        heures_out_dernier = pd.Series(0, index=heures_index, dtype=int)
        if len(df_out) > 0 and "Cle_Match" in df_out.columns:
            # Dernier OUT par Cle_Match
            idx_last_out = df_out.sort_values(["Cle_Match", "Datetime"]).groupby("Cle_Match").tail(1)
            heures_out_dernier = idx_last_out["Datetime"].dt.hour.value_counts().sort_index()
            heures_out_dernier = heures_out_dernier.reindex(heures_index, fill_value=0)
        result["heures_out_counts"] = heures_out_dernier

        # 3) Profits par heure/jour/mois (OUT uniquement)
        jours_index = pd.Index(range(7), name="Jour")
        mois_index = pd.Index(range(1, 13), name="Mois")
        if len(df_out) > 0:
            # Sommes nettes (peuvent masquer des pertes si positives)
            profits_par_heure = df_out.groupby(df_out["Datetime"].dt.hour)["Profit"].sum().sort_index().reindex(heures_index, fill_value=0.0)
            profits_par_jour = df_out.groupby(df_out["Datetime"].dt.dayofweek)["Profit"].sum().sort_index().reindex(jours_index, fill_value=0.0)
            profits_par_mois = df_out.groupby(df_out["Datetime"].dt.month)["Profit"].sum().sort_index().reindex(mois_index, fill_value=0.0)

            # Séparer correctement: somme des profits positifs uniquement et somme ABS des pertes uniquement
            df_pos = df_out[df_out["Profit"] > 0]
            df_neg = df_out[df_out["Profit"] < 0]

            profits_pos_h = df_pos.groupby(df_pos["Datetime"].dt.hour)["Profit"].sum().sort_index().reindex(heures_index, fill_value=0.0)
            pertes_abs_h = (-df_neg.groupby(df_neg["Datetime"].dt.hour)["Profit"].sum()).sort_index().reindex(heures_index, fill_value=0.0)

            profits_pos_d = df_pos.groupby(df_pos["Datetime"].dt.dayofweek)["Profit"].sum().sort_index().reindex(jours_index, fill_value=0.0)
            pertes_abs_d = (-df_neg.groupby(df_neg["Datetime"].dt.dayofweek)["Profit"].sum()).sort_index().reindex(jours_index, fill_value=0.0)

            profits_pos_m = df_pos.groupby(df_pos["Datetime"].dt.month)["Profit"].sum().sort_index().reindex(mois_index, fill_value=0.0)
            pertes_abs_m = (-df_neg.groupby(df_neg["Datetime"].dt.month)["Profit"].sum()).sort_index().reindex(mois_index, fill_value=0.0)
        else:
            profits_par_heure = pd.Series(0.0, index=heures_index, dtype=float)
            profits_par_jour = pd.Series(0.0, index=jours_index, dtype=float)
            profits_par_mois = pd.Series(0.0, index=mois_index, dtype=float)
            profits_pos_h = profits_par_heure.copy()
            pertes_abs_h = profits_par_heure.copy()
            profits_pos_d = profits_par_jour.copy()
            pertes_abs_d = profits_par_jour.copy()
            profits_pos_m = profits_par_mois.copy()
            pertes_abs_m = profits_par_mois.copy()

        result["profits_par_heure_out"] = profits_par_heure
        result["profits_par_jour_out"] = profits_par_jour
        result["profits_par_mois_out"] = profits_par_mois
        result["profits_pos_par_heure_out"] = profits_pos_h
        result["pertes_abs_par_heure_out"] = pertes_abs_h
        result["profits_pos_par_jour_out"] = profits_pos_d
        result["pertes_abs_par_jour_out"] = pertes_abs_d
        result["profits_pos_par_mois_out"] = profits_pos_m
        result["pertes_abs_par_mois_out"] = pertes_abs_m

        # Identifier best/worst heures/jours/mois
        def _best_worst(series):
            if series is None or len(series) == 0:
                return None, None
            best_idx = series.idxmax()
            worst_idx = series.idxmin()
            return int(best_idx), int(worst_idx)

        result["best_hour"], result["worst_hour"] = _best_worst(profits_par_heure)
        result["best_day"], result["worst_day"] = _best_worst(profits_par_jour)
        result["best_month"], result["worst_month"] = _best_worst(profits_par_mois)

        # 3bis) Comptes TP/SL par heure/jour/mois basés sur le dernier OUT et le profit total du trade
        tpsl_by_hour = pd.Series(0, index=heures_index, dtype=int)
        sl_by_hour = pd.Series(0, index=heures_index, dtype=int)
        tpsl_by_day = pd.Series(0, index=jours_index, dtype=int)
        sl_by_day = pd.Series(0, index=jours_index, dtype=int)
        tpsl_by_month = pd.Series(0, index=mois_index, dtype=int)
        sl_by_month = pd.Series(0, index=mois_index, dtype=int)
        if len(df_out) > 0 and "Cle_Match" in df_out.columns:
            trade_profit = df.groupby("Cle_Match")["Profit"].sum().reset_index()
            last_out = df_out.sort_values(["Cle_Match", "Datetime"]).groupby("Cle_Match").tail(1)[["Cle_Match", "Datetime"]]
            trades_final = trade_profit.merge(last_out, on="Cle_Match", how="inner")
            trades_final["hour"] = trades_final["Datetime"].dt.hour
            trades_final["day"] = trades_final["Datetime"].dt.dayofweek
            trades_final["month"] = trades_final["Datetime"].dt.month
            tps = trades_final[trades_final["Profit"] > 0]
            sls = trades_final[trades_final["Profit"] < 0]
            tpsl_by_hour = tps.groupby("hour").size().reindex(heures_index, fill_value=0)
            sl_by_hour = sls.groupby("hour").size().reindex(heures_index, fill_value=0)
            tpsl_by_day = tps.groupby("day").size().reindex(jours_index, fill_value=0)
            sl_by_day = sls.groupby("day").size().reindex(jours_index, fill_value=0)
            tpsl_by_month = tps.groupby("month").size().reindex(mois_index, fill_value=0)
            sl_by_month = sls.groupby("month").size().reindex(mois_index, fill_value=0)
        result["tp_par_heure"] = tpsl_by_hour
        result["sl_par_heure"] = sl_by_hour
        result["tp_par_jour"] = tpsl_by_day
        result["sl_par_jour"] = sl_by_day
        result["tp_par_mois"] = tpsl_by_month
        result["sl_par_mois"] = sl_by_month

        # 4) Évolution cumulée (linéaire) des profits par trade au moment du dernier OUT
        cumul_df = pd.DataFrame(columns=["Datetime", "Profit_Trade", "Cumul_Profit"]) 
        if "Cle_Match" in df.columns and len(df_out) > 0:
            trades_sum = df.groupby("Cle_Match")["Profit"].sum().reset_index()
            last_out = df_out.sort_values(["Cle_Match", "Datetime"]).groupby("Cle_Match").tail(1)[["Cle_Match", "Datetime"]]
            series_trades = trades_sum.merge(last_out, on="Cle_Match", how="inner").dropna(subset=["Datetime"]) 
            series_trades = series_trades.sort_values("Datetime")
            series_trades["Cumul_Profit"] = series_trades["Profit"].cumsum()
            cumul_df = series_trades.rename(columns={"Profit": "Profit_Trade"})[["Datetime", "Profit_Trade", "Cumul_Profit"]]
        result["cumul_evolution"] = cumul_df

        # 5) Durée moyenne/médiane des trades (IN -> dernier OUT)
        duree_moyenne_minutes = None
        duree_mediane_minutes = None
        if "Cle_Match" in df.columns and len(df_in) > 0 and len(df_out) > 0:
            first_in = df_in.sort_values(["Cle_Match", "Datetime"]).groupby("Cle_Match").head(1)[["Cle_Match", "Datetime"]].rename(columns={"Datetime": "InTime"})
            last_out = df_out.sort_values(["Cle_Match", "Datetime"]).groupby("Cle_Match").tail(1)[["Cle_Match", "Datetime"]].rename(columns={"Datetime": "OutTime"})
            joined = first_in.merge(last_out, on="Cle_Match", how="inner")
            if len(joined) > 0:
                durees = (joined["OutTime"] - joined["InTime"]).dt.total_seconds() / 60.0
                duree_moyenne_minutes = float(durees.mean())
                duree_mediane_minutes = float(durees.median())
        result["duree_moyenne_minutes"] = duree_moyenne_minutes
        result["duree_mediane_minutes"] = duree_mediane_minutes

        # Préparer les données d'évolution pour le web
        evolution_somme_cumulee = []
        print(f"[DEBUG] Colonnes disponibles pour evolution_somme_cumulee: {list(df.columns)}")
        
        # Chercher les colonnes de date et solde avec différentes variantes
        date_col = None
        solde_col = None
        
        for col in df.columns:
            if col.lower() in ['datetime', 'date', 'timestamp']:
                date_col = col
            elif col.lower() in ['solde_cumule', 'solde_cumulé', 'solde_cumulee', 'cumul', 'balance']:
                solde_col = col
        
        print(f"[DEBUG] Colonne date trouvée: {date_col}, Colonne solde trouvée: {solde_col}")
        
        if date_col and solde_col:
            for _, row in df.iterrows():
                try:
                    date_val = row[date_col]
                    solde_val = row[solde_col]
                    
                    # Convertir la date en format ISO si possible
                    if hasattr(date_val, 'isoformat'):
                        date_str = date_val.isoformat()
                    else:
                        date_str = str(date_val)
                    
                    evolution_somme_cumulee.append({
                        'date': date_str,
                        'solde': float(round(float(solde_val), 2))
                    })
                except Exception as e:
                    print(f"[WARNING] Erreur lors du traitement de la ligne: {e}")
                    continue
        
        print(f"[DEBUG] Nombre de points d'évolution créés: {len(evolution_somme_cumulee)}")
        result["evolution_somme_cumulee"] = evolution_somme_cumulee

        return result

    def calculer_performance_par_session(self, df: pd.DataFrame):
        """Calcule la performance par session (Asie/Europe/Amérique).

        Règles:
        - Session Asie: 00:00-07:59 UTC
        - Session Europe: 08:00-15:59 UTC
        - Session Amérique: 16:00-23:59 UTC

        Métriques:
        - Côté IN (qualité d'ouverture): nb IN, taux de réussite des trades ouverts dans la session
        - Côté OUT (attribution du PnL): PnL total basé sur le DERNIER OUT (profit total du trade), nb TP/SL
        - Décliné par paire (Symbole_ordre) et au total
        """
        result = {
            "sessions_total": {},
            "sessions_par_pair": {}
        }

        if "Heure d'ouverture" not in df.columns or "Direction" not in df.columns:
            return result

        data = df.copy()
        data["Datetime"] = pd.to_datetime(data["Heure d'ouverture"], errors='coerce')
        data = data[data["Datetime"].notna()]
        if len(data) == 0:
            return result

        # Helper: assign session by hour
        def heure_to_session(h):
            try:
                h = int(h)
            except Exception:
                return "Autre"
            if 0 <= h <= 7:
                return "Asie"
            if 8 <= h <= 15:
                return "Europe"
            if 16 <= h <= 23:
                return "Amérique"
            return "Autre"

        data["Session"] = data["Datetime"].dt.hour.apply(heure_to_session)

        # Calcul par paire puis total
        symbols = list(data["Symbole_ordre"].dropna().unique()) if "Symbole_ordre" in data.columns else ["TOTAL"]
        if "Symbole_ordre" not in data.columns:
            data["Symbole_ordre"] = "TOTAL"

        for symbole in symbols:
            d = data[data["Symbole_ordre"] == symbole].copy()
            if len(d) == 0:
                continue

            # Vue IN: qualité des ouvertures par session
            d_in = d[d["Direction"] == "in"].copy()
            in_by_session = d_in.groupby("Session").size().reindex(["Asie","Europe","Amérique"], fill_value=0)

            # Pour le taux de réussite côté IN, on a besoin du résultat du trade complet
            taux_reussite_in = {"Asie": 0.0, "Europe": 0.0, "Amérique": 0.0}
            if "Cle_Match" in d.columns and len(d_in) > 0:
                # Profit total par trade
                profit_par_trade = d.groupby("Cle_Match")["Profit"].sum()
                # Session d'ouverture du trade (session du premier IN)
                first_in = d_in.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").head(1)[["Cle_Match","Session"]]
                first_in = first_in.dropna(subset=["Cle_Match"]).set_index("Cle_Match")
                joined = first_in.join(profit_par_trade, how="left").rename(columns={"Profit":"Profit_Trade"})
                for sess in ["Asie","Europe","Amérique"]:
                    subset = joined[joined["Session"] == sess]
                    denom = len(subset)
                    if denom > 0:
                        taux = (len(subset[subset["Profit_Trade"] > 0]) / denom) * 100.0
                        taux_reussite_in[sess] = round(float(taux), 2)

            # Vue OUT: attribution PnL sur le DERNIER OUT et comptage TP/SL par session de sortie
            d_out = d[d["Direction"] == "out"].copy()
            pnl_session = {"Asie": 0.0, "Europe": 0.0, "Amérique": 0.0}
            tp_session = {"Asie": 0, "Europe": 0, "Amérique": 0}
            sl_session = {"Asie": 0, "Europe": 0, "Amérique": 0}
            if len(d_out) > 0 and "Cle_Match" in d_out.columns:
                # Profit total du trade + session du dernier OUT
                trade_profit = d.groupby("Cle_Match")["Profit"].sum().reset_index()
                last_out = d_out.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").tail(1)[["Cle_Match","Session"]]
                final = trade_profit.merge(last_out, on="Cle_Match", how="inner")
                for sess in ["Asie","Europe","Amérique"]:
                    pnl = final[final["Session"] == sess]["Profit"].sum()
                    pnl_session[sess] = float(round(pnl, 2))
                    tp_session[sess] = int((final[(final["Session"] == sess) & (final["Profit"] > 0)]).shape[0])
                    sl_session[sess] = int((final[(final["Session"] == sess) & (final["Profit"] < 0)]).shape[0])

            bloc = {
                "in_count": {k: int(in_by_session.get(k, 0)) for k in ["Asie","Europe","Amérique"]},
                "taux_reussite_in_pct": taux_reussite_in,
                "pnl_out": pnl_session,
                "tp_out": tp_session,
                "sl_out": sl_session
            }

            if symbole == "TOTAL":
                result["sessions_total"] = bloc
            else:
                result["sessions_par_pair"][symbole] = bloc

        # Si on a plusieurs paires, calculer aussi un total global en agrégeant toutes les lignes
        if "TOTAL" not in symbols and "Symbole_ordre" in df.columns:
            result_global = self.calculer_performance_par_session(data.assign(Symbole_ordre="TOTAL"))
            result["sessions_total"] = result_global.get("sessions_total", {})

        return result

    def calculer_patterns(self, df: pd.DataFrame, min_support: float = 0.03, min_confidence: float = 0.55, top_k: int = 10, n_permutations: int = 200, max_itemset_size: int = 3):
        """Détecte des patterns (itemsets 1-2) et génère des règles vers TP/SL avec p-values.

        Contraintes d'items (plus intuitives):
        - Durée du trade (IN -> dernier OUT): buckets {D<30m, D30-120m, >120m}
        - Sens (buy/sell)
        - Heure d'ouverture (buckets {H[0-7], H[8-11], H[12-15], H[16-19], H[20-23]})
        - Session d'ouverture {Asie, Europe, Amérique}

        Calculs retournés pour chaque règle Items ⇒ TP (ou SL):
        - count, support, confidence, lift, p_value (test permutation sur la confidence)
        """
        results = {"top_tp": [], "top_sl": []}
        if "Cle_Match" not in df.columns or len(df) == 0:
            return results

        data = df.copy()
        data["Datetime"] = pd.to_datetime(data.get("Heure d'ouverture"), errors='coerce')
        data = data[data["Datetime"].notna()]

        # Premier IN et dernier OUT par trade
        df_in = data[data["Direction"] == "in"].copy()
        df_out = data[data["Direction"] == "out"].copy()
        if len(df_in) == 0 or len(df_out) == 0:
            return results

        first_in = df_in.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").head(1)
        last_out = df_out.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").tail(1)
        profit_par_trade = data.groupby("Cle_Match")["Profit"].sum()
        pips_par_trade = data.groupby("Cle_Match")["Profit_pips"].sum() if "Profit_pips" in data.columns else None
        outs_count = df_out.groupby("Cle_Match").size()

        # Construction du tableau des trades complets
        trades = first_in[["Cle_Match","Symbole_ordre","Type_ordre","Datetime"]].rename(columns={"Datetime":"InTime"}).copy()
        trades = trades.merge(last_out[["Cle_Match","Datetime"]].rename(columns={"Datetime":"OutTime"}), on="Cle_Match", how="inner")
        trades["Profit_Trade"] = trades["Cle_Match"].map(profit_par_trade)
        if pips_par_trade is not None:
            trades["Pips_Trade"] = trades["Cle_Match"].map(pips_par_trade)
        trades["Nb_OUTs"] = trades["Cle_Match"].map(outs_count).fillna(0).astype(int)

        # Sessions et features dérivées
        def heure_to_session(h):
            if 0 <= h <= 7:
                return "Asie"
            if 8 <= h <= 15:
                return "Europe"
            return "Amérique"

        trades["Session_IN"] = trades["InTime"].dt.hour.apply(heure_to_session)
        trades["Heure_IN"] = trades["InTime"].dt.hour
        trades["Duree_min"] = (trades["OutTime"] - trades["InTime"]).dt.total_seconds() / 60.0

        def bucket_hour(h):
            try:
                h = int(h)
            except Exception:
                return "HNA"
            if h < 8:
                return "H[0-7]"
            if h < 12:
                return "H[8-11]"
            if h < 16:
                return "H[12-15]"
            if h < 20:
                return "H[16-19]"
            return "H[20-23]"

        def bucket_duration(m):
            try:
                m = float(m)
            except Exception:
                return "DUR_NA"
            if m < 30:
                return "D<30m"
            if m <= 120:
                return "D30-120m"
            if m <= 360:  # 2h-6h
                return "D2-6h"
            if m <= 720:  # 6h-12h
                return "D6-12h"
            if m <= 1440:  # 12h-24h
                return "D12-24h"
            return ">24h"

        def bucket_outs(n):
            try:
                n = int(n)
            except Exception:
                return "OUT_NA"
            if n <= 1:
                return "OUT=1"
            if n <= 3:
                return "OUT=2-3"
            return "OUT>=4"

        trades["Heure_Bucket"] = trades["Heure_IN"].apply(bucket_hour)
        trades["Duree_Bucket"] = trades["Duree_min"].apply(bucket_duration)
        # NOTE: on n'utilise plus OUTS_Bucket dans les items (simplification demandée)
        trades["OUTS_Bucket"] = trades["Nb_OUTs"].apply(bucket_outs)
        trades["Result"] = trades["Profit_Trade"].apply(lambda x: "TP" if x > 0 else ("SL" if x < 0 else "NEUTRE"))

        # Ensemble d'items
        def items_for_row(r):
            items = set()
            # Seuls les 4 attributs demandés
            if pd.notna(r.get("Type_ordre")):
                items.add(f"DIR={str(r['Type_ordre']).lower()}")
            items.add(f"SESSION={r['Session_IN']}")
            items.add(f"{r['Heure_Bucket']}")
            items.add(f"{r['Duree_Bucket']}")
            return items

        trades["Items"] = trades.apply(items_for_row, axis=1)

        # Comptages
        N = len(trades)
        base_tp = (trades["Result"] == "TP").mean() if N > 0 else 0
        base_sl = (trades["Result"] == "SL").mean() if N > 0 else 0

        from collections import Counter
        count_1 = Counter()
        count_1_tp = Counter()
        count_1_sl = Counter()
        count_2 = Counter()
        count_2_tp = Counter()
        count_2_sl = Counter()
        count_3 = Counter()
        count_3_tp = Counter()
        count_3_sl = Counter()

        for _, r in trades.iterrows():
            items = sorted(list(r["Items"]))
            res = r["Result"]
            # size-1
            for a in items:
                count_1[a] += 1
                if res == "TP":
                    count_1_tp[a] += 1
                elif res == "SL":
                    count_1_sl[a] += 1
            # size-2 (combinatoire limitée)
            for i in range(len(items)):
                for j in range(i+1, len(items)):
                    pair = (items[i], items[j])
                    count_2[pair] += 1
                    if res == "TP":
                        count_2_tp[pair] += 1
                    elif res == "SL":
                        count_2_sl[pair] += 1
            # size-3 (au plus 4 items donc ≤ 4 combinaisons par trade)
            if max_itemset_size >= 3 and len(items) >= 3:
                for i in range(len(items)):
                    for j in range(i+1, len(items)):
                        for k in range(j+1, len(items)):
                            trip = (items[i], items[j], items[k])
                            count_3[trip] += 1
                            if res == "TP":
                                count_3_tp[trip] += 1
                            elif res == "SL":
                                count_3_sl[trip] += 1

        def build_rows(counter_all, counter_pos, baseline):
            rows = []
            for item, cnt in counter_all.items():
                support = cnt / N
                if support < min_support:
                    continue
                pos = counter_pos.get(item, 0)
                conf = pos / cnt if cnt > 0 else 0
                lift = (conf / baseline) if baseline > 0 else 0
                rows.append({
                    "items": item if isinstance(item, str) else " & ".join(item),
                    "count": int(cnt),
                    "support": round(support, 3),
                    "confidence": round(conf, 3),
                    "lift": round(lift, 3)
                })
            # Filtrer min_confidence et trier par lift puis support
            rows = [r for r in rows if r["confidence"] >= min_confidence]
            rows.sort(key=lambda x: (x["lift"], x["support"]), reverse=True)
            return rows[:top_k]

        top_tp = build_rows(count_1, count_1_tp, base_tp) + build_rows(count_2, count_2_tp, base_tp)
        top_sl = build_rows(count_1, count_1_sl, base_sl) + build_rows(count_2, count_2_sl, base_sl)
        if max_itemset_size >= 3:
            top_tp += build_rows(count_3, count_3_tp, base_tp)
            top_sl += build_rows(count_3, count_3_sl, base_sl)

        # Garder top_k au global après concat (tri provisoire, la p-value sera ajoutée ensuite)
        top_tp = sorted(top_tp, key=lambda x: (x["lift"], x["support"]), reverse=True)[: top_k * 3]
        top_sl = sorted(top_sl, key=lambda x: (x["lift"], x["support"]), reverse=True)[: top_k * 3]

        # === p-values par permutation sur la confidence ===
        import numpy as np

        def compute_confidence_for_items(items_set, target):
            mask = trades["Items"].apply(lambda s: items_set.issubset(s))
            cnt = int(mask.sum())
            if cnt == 0:
                return 0.0, cnt
            conf = float((trades.loc[mask, "Result"] == target).mean())
            return conf, cnt

        def permutation_p_value(items_str, target, observed_conf):
            # Reconstruire l'ensemble d'items
            items_set = set([s.strip() for s in items_str.split("&")]) if "&" in items_str else {items_str}
            items_set = {s.strip() for s in items_set}
            mask = trades["Items"].apply(lambda s: items_set.issubset(s))
            cnt = int(mask.sum())
            if cnt == 0:
                return 1.0
            y = trades["Result"].values
            idx = np.where(mask.values)[0]
            if len(idx) == 0:
                return 1.0
            successes = (y[idx] == target).sum()
            # Si peu d'observations, garder p=1 par prudence
            if cnt < 10:
                return 1.0
            more_extreme = 0
            for _ in range(int(max(10, n_permutations))):
                y_perm = np.random.permutation(y)
                conf_perm = (y_perm[idx] == target).mean()
                if conf_perm >= observed_conf:
                    more_extreme += 1
            pval = (more_extreme + 1) / (n_permutations + 1)
            return float(round(pval, 4))

        for row in top_tp:
            row["p_value"] = permutation_p_value(row["items"], "TP", row["confidence"])
        for row in top_sl:
            row["p_value"] = permutation_p_value(row["items"], "SL", row["confidence"])

        # Correction des tests multiples (FDR Benjamini–Hochberg) et nouveau classement
        def fdr_bh(rows):
            ps = [r.get("p_value", 1.0) for r in rows]
            m = max(len(ps), 1)
            order = np.argsort(ps)
            qvals = [1.0] * len(ps)
            min_q = 1.0
            for rank, idx in enumerate(order, start=1):
                p = ps[idx]
                q = p * m / rank
                if q < min_q:
                    min_q = q
                qvals[idx] = min_q
            for i, r in enumerate(rows):
                r["q_value"] = float(round(min(qvals[i], 1.0), 4))
            return rows

        top_tp = fdr_bh(top_tp)
        top_sl = fdr_bh(top_sl)

        # Reclasser: p-value croissante, puis q-value, puis lift décroissant, support décroissant
        top_tp = sorted(top_tp, key=lambda x: (x.get("p_value", 1.0), x.get("q_value", 1.0), -x["lift"], -x["support"]))[:top_k]
        top_sl = sorted(top_sl, key=lambda x: (x.get("p_value", 1.0), x.get("q_value", 1.0), -x["lift"], -x["support"]))[:top_k]

        results["top_tp"] = top_tp
        results["top_sl"] = top_sl
        return results

    def calculer_modele_influence(self, df: pd.DataFrame, include_interactions: bool = True, top_k: int = 20):
        """Modèle logistique sans a priori pour TP (1) vs SL (0) avec interactions.

        Retourne un DataFrame avec colonnes: feature, coef, odds_ratio, p_value.
        """
        try:
            import numpy as np
            import pandas as pd
            import statsmodels.api as sm
            backend = "statsmodels"
        except Exception as e:
            print(f"[WARNING] statsmodels not available: {e}")
            backend = "sklearn"

        if "Cle_Match" not in df.columns or len(df) == 0:
            return pd.DataFrame()

        data = df.copy()
        data["Datetime"] = pd.to_datetime(data.get("Heure d'ouverture"), errors='coerce')
        data = data[data["Datetime"].notna()]
        df_in = data[data["Direction"] == "in"].copy()
        df_out = data[data["Direction"] == "out"].copy()
        if len(df_in) == 0 or len(df_out) == 0:
            return pd.DataFrame()

        first_in = df_in.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").head(1)
        last_out = df_out.sort_values(["Cle_Match","Datetime"]).groupby("Cle_Match").tail(1)
        profit_par_trade = data.groupby("Cle_Match")["Profit"].sum()

        trades = first_in[["Cle_Match","Type_ordre","Datetime"]].rename(columns={"Datetime":"InTime"}).copy()
        trades = trades.merge(last_out[["Cle_Match","Datetime"]].rename(columns={"Datetime":"OutTime"}), on="Cle_Match", how="inner")
        trades["Profit_Trade"] = trades["Cle_Match"].map(profit_par_trade)

        def heure_to_session(h):
            if 0 <= h <= 7:
                return "Asie"
            if 8 <= h <= 15:
                return "Europe"
            return "Amérique"

        trades["Session_IN"] = trades["InTime"].dt.hour.apply(heure_to_session)
        trades["Heure_IN"] = trades["InTime"].dt.hour
        trades["Duree_min"] = (trades["OutTime"] - trades["InTime"]).dt.total_seconds() / 60.0

        def bucket_hour(h):
            try:
                h = int(h)
            except Exception:
                return "HNA"
            if h < 8:
                return "H[0-7]"
            if h < 12:
                return "H[8-11]"
            if h < 16:
                return "H[12-15]"
            if h < 20:
                return "H[16-19]"
            return "H[20-23]"

        def bucket_duration(m):
            try:
                m = float(m)
            except Exception:
                return "DUR_NA"
            if m < 30:
                return "D<30m"
            if m <= 120:
                return "D30-120m"
            if m <= 360:
                return "D2-6h"
            if m <= 720:
                return "D6-12h"
            if m <= 1440:
                return "D12-24h"
            return ">24h"

        trades["Heure_Bucket"] = trades["Heure_IN"].apply(bucket_hour)
        trades["Duree_Bucket"] = trades["Duree_min"].apply(bucket_duration)
        trades["y"] = trades["Profit_Trade"].apply(lambda x: 1 if x > 0 else (0 if x < 0 else np.nan))
        trades = trades[trades["y"].notna()]

        X = pd.get_dummies(trades[["Type_ordre","Session_IN","Heure_Bucket","Duree_Bucket"]], drop_first=True)
        if include_interactions:
            # Interactions pairwise: créer manuellement pour lisibilité
            cols = list(X.columns)
            for i in range(len(cols)):
                for j in range(i+1, len(cols)):
                    X[f"{cols[i]}*{cols[j]}"] = X[cols[i]] * X[cols[j]]

        y = trades["y"].astype(int)
        if backend == "statsmodels":
            X_sm = sm.add_constant(X, has_constant='add')
            try:
                model = sm.Logit(y, X_sm).fit(disp=False, maxiter=200)
            except Exception as e:
                print(f"[WARNING] Logit failed: {e}")
                return pd.DataFrame()

            params = model.params
            pvalues = model.pvalues
            oratios = np.exp(params)
            res_df = pd.DataFrame({
                "feature": params.index,
                "coef": params.values,
                "odds_ratio": oratios.values,
                "p_value": pvalues.values
            })
            res_df = res_df[res_df["feature"] != "const"].copy()
            res_df["odds_ratio"] = res_df["odds_ratio"].round(3)
            res_df["coef"] = res_df["coef"].round(4)
            res_df["p_value"] = res_df["p_value"].round(4)
            res_df = res_df.sort_values(["p_value", "odds_ratio"], ascending=[True, False]).head(top_k)
            return res_df
        else:
            try:
                from sklearn.linear_model import LogisticRegression
            except Exception as e:
                print(f"[WARNING] sklearn not available: {e}")
                # Fallback sans dépendances: tests bi-variés par variable (2-proportions z-test)
                import math
                rows = []
                for col in X.columns:
                    try:
                        mask1 = X[col] == 1
                        mask0 = X[col] == 0
                        n1 = int(mask1.sum())
                        n0 = int(mask0.sum())
                        if n1 == 0 or n0 == 0:
                            continue
                        y1 = int(y[mask1].sum())
                        y0 = int(y[mask0].sum())
                        p1 = y1 / n1
                        p0 = y0 / n0
                        p_pool = (y1 + y0) / (n1 + n0)
                        se = math.sqrt(max(p_pool * (1 - p_pool) * (1 / n1 + 1 / n0), 1e-12))
                        z = (p1 - p0) / se if se > 0 else 0.0
                        # p-value bilatérale via erf
                        cdf = 0.5 * (1.0 + math.erf(abs(z) / math.sqrt(2)))
                        pval = float(round(2 * (1 - cdf), 4))
                        # Odds ratio avec lissage 0.5 (Haldane-Anscombe)
                        a = y1 + 0.5
                        b = (n1 - y1) + 0.5
                        c = y0 + 0.5
                        d = (n0 - y0) + 0.5
                        oratio = (a / b) / (c / d)
                        rows.append({
                            "feature": col,
                            "coef": round(p1 - p0, 4),
                            "odds_ratio": round(oratio, 3),
                            "p_value": pval
                        })
                    except Exception:
                        continue
                if not rows:
                    return pd.DataFrame()
                res_df = pd.DataFrame(rows).sort_values(["p_value", "odds_ratio"], ascending=[True, False]).head(top_k)
                return res_df

            model = LogisticRegression(max_iter=200, solver="liblinear")
            try:
                model.fit(X, y)
            except Exception as e:
                print(f"[WARNING] Sklearn logistic fit failed: {e}")
                return pd.DataFrame()
            coefs = model.coef_[0]
            oratios = np.exp(coefs)
            res_df = pd.DataFrame({
                "feature": X.columns,
                "coef": coefs,
                "odds_ratio": oratios,
                "p_value": [None] * len(coefs)
            })
            res_df["odds_ratio"] = res_df["odds_ratio"].round(3)
            res_df["coef"] = res_df["coef"].round(4)
            # Trier par importance absolue du coef
            res_df = res_df.reindex(res_df["coef"].abs().sort_values(ascending=False).index)
            res_df = res_df.head(top_k)
            return res_df
    
    def create_excel_report(self, df_final, reports_folder, timestamp, filter_type=None):
        """Crée un rapport Excel complet avec graphiques"""
        try:
            print(f"[DEBUG] Starting Excel report creation")
            
            # Calculer les statistiques avancees
            stats_avancees = self.calculer_statistiques_avancees(df_final)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # === ONGLET 1: RÉSUMÉ GLOBAL ===
            ws_resume = wb.create_sheet("📊 Résumé Global")
            
            # Titre principal
            ws_resume.merge_cells('A1:H1')
            cell_titre = ws_resume['A1']
            titre_type = "FOREX" if filter_type == 'forex' else "AUTRES INSTRUMENTS" if filter_type == 'autres' else "TOUS INSTRUMENTS"
            cell_titre.value = f"RAPPORT {titre_type} - {datetime.now().strftime('%d/%m/%Y')}"
            cell_titre.font = Font(size=16, bold=True, color="FFFFFF")
            cell_titre.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_titre.alignment = Alignment(horizontal="center", vertical="center")
            
            # Statistiques globales basées sur les trades complets
            total_trades = stats_avancees["total_trades_complets"]
            trades_gagnants = stats_avancees["trades_gagnants_complets"]
            trades_perdants = stats_avancees["trades_perdants_complets"]
            trades_neutres = stats_avancees["trades_neutres_complets"]
            trades_avec_resultat = trades_gagnants + trades_perdants
            
            # Calculer le profit total basé sur les trades complets
            trades_complets = df_final.groupby("Cle_Match").agg({
                "Profit": "sum"
            }).reset_index()
            profit_total_lineaire = trades_complets["Profit"].sum()
            
            profit_total_compose = df_final['Profit_cumule'].iloc[-1] if len(df_final) > 0 else 0

            # Pips totaux et statistiques de pips calculés PAR TRADE COMPLET
            trades_complets_pips = df_final.groupby("Cle_Match").agg({
                "Profit_pips": "sum"
            }).reset_index()
            pips_totaux = trades_complets_pips["Profit_pips"].sum() if len(trades_complets_pips) > 0 else 0
            pips_moyen_par_trade = (pips_totaux / total_trades) if total_trades > 0 else 0
            pips_pertes_total = abs(trades_complets_pips[trades_complets_pips["Profit_pips"] < 0]["Profit_pips"].sum()) if len(trades_complets_pips) > 0 else 0
            pips_moyen_pertes = abs(trades_complets_pips[trades_complets_pips["Profit_pips"] < 0]["Profit_pips"].mean()) if len(trades_complets_pips[trades_complets_pips["Profit_pips"] < 0]) > 0 else 0
            solde_final = df_final['Solde_cumule'].iloc[-1] if len(df_final) > 0 else self.solde_initial
            rendement_pct = ((solde_final - self.solde_initial) / self.solde_initial * 100)
            
            difference_compose = profit_total_compose - profit_total_lineaire
            gain_compose_pct = ((profit_total_compose / profit_total_lineaire - 1) * 100) if profit_total_lineaire != 0 else 0
            taux_reussite = (trades_gagnants / trades_avec_resultat * 100) if trades_avec_resultat > 0 else 0
            
            # Tableau des statistiques
            stats_data = [
                ["📊 STATISTIQUES PRINCIPALES", ""],
                ["", ""],
                ["💰 Solde initial", f"{self.solde_initial:,.2f} €"],
                ["💳 Solde final (composé)", f"{solde_final:,.2f} €"],
                ["📈 Profit total (linéaire)", f"{profit_total_lineaire:,.2f} €"],
                ["🚀 Profit total (composé)", f"{profit_total_compose:,.2f} €"],
                ["⚡ Profits liés aux intérêts composés", f"{difference_compose:,.2f} € (+{gain_compose_pct:.2f}%)"],
                ["📊 Rendement global", f"{rendement_pct:.2f} %"],
                ["🎯 Pips/Points totaux", f"{pips_totaux:,.2f}"],
                ["🎯 Pips moyens par trade", f"{pips_moyen_par_trade:,.2f}"],
                ["❌ Pips perdus (total)", f"{pips_pertes_total:,.2f}"],
                ["❌ Pips moyens lors des pertes", f"{pips_moyen_pertes:,.2f}"],
                ["", ""],
                ["📉 ANALYSE DU DRAWDOWN", ""],
                ["", ""],
                ["📉 Drawdown maximum", f"{stats_avancees['drawdown_max_pct']:.2f} %"],
                ["💸 Drawdown max (euros)", f"{stats_avancees['drawdown_max_euros']:,.2f} €"],
                ["", ""],
                ["🔢 ANALYSE DES TRADES (hors neutres)", ""],
                ["", ""],
                ["📈 Total trades", total_trades],
                ["✅ Trades gagnants", trades_gagnants],
                ["❌ Trades perdants", trades_perdants],
                ["⚪ Trades neutres (exclus)", f"{trades_neutres} (non comptés)"],
                ["🎯 Taux de réussite", f"{taux_reussite:.1f} % (sur {trades_avec_resultat} trades)"],
                ["", ""],
                ["📊 DÉTAIL DES OPÉRATIONS", ""],
                ["", ""],
                ["📈 Total opérations", f"{len(df_final)} (toutes les lignes)"],
                ["📈 Total trades complets", f"{total_trades} (1 IN + 1 ou plusieurs OUT)"],
                ["📈 Trades avec management dynamique", f"{len(df_final[df_final['Direction'] == 'out']) - total_trades} opérations partielles"],
                ["", ""],
                ["📈 SÉRIES ET MOYENNES", ""],
                ["", ""],
                ["🔥 Gains consécutifs max", f"{stats_avancees['gains_consecutifs_max']} trades"],
                ["💔 Pertes consécutives max", f"{stats_avancees['pertes_consecutives_max']} trades"],
                ["💚 Gain moyen", f"{stats_avancees['gain_moyen']:,.2f} €"],
                ["💔 Perte moyenne", f"{stats_avancees['perte_moyenne']:,.2f} €"],
            ]
            
            # Ajout des statistiques par fichier si disponibles
            if self.statistiques_fichiers:
                stats_data.extend([
                    ["", ""],
                    ["📁 DÉTAIL PAR FICHIER", ""],
                    ["", ""]
                ])
                for fichier, stats in self.statistiques_fichiers.items():
                    stats_data.append([f"📄 {fichier[:30]}...", f"{stats['trades']} trades complets, {stats['exclus']} exclus"])
            
            for row_idx, (label, value) in enumerate(stats_data, 3):
                ws_resume[f'A{row_idx}'] = label
                ws_resume[f'B{row_idx}'] = value
                
                # Formatage des en-têtes
                if any(word in label for word in ["STATISTIQUES", "ANALYSE", "DRAWDOWN", "SÉRIES", "DÉTAIL"]):
                    ws_resume[f'A{row_idx}'].font = Font(bold=True, color="366092")
                    ws_resume[f'A{row_idx}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                ws_resume[f'A{row_idx}'].alignment = Alignment(horizontal="left")
                ws_resume[f'B{row_idx}'].alignment = Alignment(horizontal="right")
            
            print(f"[DEBUG] Summary sheet created with {len(stats_data)} rows")

            # === AGRÉGATIONS POUR GRAPHIQUES (placer dans une feuille dédiée) ===
            aggs = self.calculer_agregations_graphes(df_final)
            sessions = self.calculer_performance_par_session(df_final)
            ws_charts = wb.create_sheet("📈 Graphiques Résumé")

            # === SECTION 1: DURÉE MOYENNE ===
            duree_m = aggs.get("duree_moyenne_minutes")
            duree_med = aggs.get("duree_mediane_minutes")
            ws_charts['A1'] = "⏱️ TEMPS MOYEN DES TRADES"
            ws_charts['A1'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A2'] = "Temps moyen (IN -> dernier OUT):"
            ws_charts['A2'].font = Font(bold=True)
            if duree_m is not None:
                heures = int(duree_m // 60)
                minutes = int(duree_m % 60)
                ws_charts['B2'] = f"{heures}h {minutes}m ({duree_m:.1f} minutes)"
            if duree_med is not None:
                ws_charts['A3'] = "Temps médian:"
                ws_charts['A3'].font = Font(bold=True)
                heures_med = int(duree_med // 60)
                minutes_med = int(duree_med % 60)
                ws_charts['B3'] = f"{heures_med}h {minutes_med}m ({duree_med:.1f} minutes)"

            # === SECTION 2: RÉPARTITION DES OUVERTURES (IN) ===
            ws_charts['A8'] = "📊 RÉPARTITION DES OUVERTURES (IN) PAR HEURE"
            ws_charts['A8'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A10'] = "Heure"
            ws_charts['B10'] = "IN (comptage)"
            heures = list(range(24))
            row_ptr = 11
            heures_in_series = aggs.get("heures_in_counts", pd.Series(dtype=int))
            for h in heures:
                ws_charts.cell(row=row_ptr, column=1, value=h)
                ws_charts.cell(row=row_ptr, column=2, value=int(heures_in_series.get(h, 0)))
                row_ptr += 1

            chart_in = BarChart()
            chart_in.title = "Répartition des ouvertures (IN) par heure"
            chart_in.y_axis.title = "Comptage"
            chart_in.x_axis.title = "Heure (0-23)"
            data_in = Reference(ws_charts, min_col=2, min_row=10, max_row=row_ptr-1)
            cats_in = Reference(ws_charts, min_col=1, min_row=11, max_row=row_ptr-1)
            chart_in.add_data(data_in, titles_from_data=True)
            chart_in.set_categories(cats_in)
            chart_in.legend.position = 'b'
            chart_in.height = 12
            chart_in.width = 25
            ws_charts.add_chart(chart_in, "E8")

            # === SECTION 3: RÉPARTITION DES FERMETURES (OUT) ===
            ws_charts['A40'] = "📊 RÉPARTITION DES FERMETURES (OUT) PAR HEURE"
            ws_charts['A40'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A42'] = "Heure"
            ws_charts['B42'] = "OUT dernier (comptage)"
            row_ptr2 = 43
            heures_out_series = aggs.get("heures_out_counts", pd.Series(dtype=int))
            for h in heures:
                ws_charts.cell(row=row_ptr2, column=1, value=h)
                ws_charts.cell(row=row_ptr2, column=2, value=int(heures_out_series.get(h, 0)))
                row_ptr2 += 1

            chart_out = BarChart()
            chart_out.title = "Répartition des fermetures (dernier OUT) par heure"
            chart_out.y_axis.title = "Comptage"
            chart_out.x_axis.title = "Heure (0-23)"
            data_out = Reference(ws_charts, min_col=2, min_row=42, max_row=row_ptr2-1)
            cats_out = Reference(ws_charts, min_col=1, min_row=43, max_row=row_ptr2-1)
            chart_out.add_data(data_out, titles_from_data=True)
            chart_out.set_categories(cats_out)
            chart_out.legend.position = 'b'
            chart_out.height = 12
            chart_out.width = 25
            ws_charts.add_chart(chart_out, "E40")

            # === SECTION 4: PROFITS PAR HEURE ===
            ws_charts['A72'] = "💰 PROFITS PAR HEURE (OUT)"
            ws_charts['A72'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A74'] = "Heure"
            ws_charts['B74'] = "Profit (€)"
            row_ptr3 = 75
            profits_par_heure = aggs.get("profits_par_heure_out", pd.Series(dtype=float))
            for h in heures:
                ws_charts.cell(row=row_ptr3, column=1, value=h)
                ws_charts.cell(row=row_ptr3, column=2, value=float(round(profits_par_heure.get(h, 0.0), 2)))
                row_ptr3 += 1
            chart_ph = BarChart()
            chart_ph.title = "Profits par heure (OUT)"
            chart_ph.y_axis.title = "Profit (€)"
            chart_ph.x_axis.title = "Heure (0-23)"
            data = Reference(ws_charts, min_col=2, min_row=74, max_row=row_ptr3-1)
            cats = Reference(ws_charts, min_col=1, min_row=75, max_row=row_ptr3-1)
            chart_ph.add_data(data, titles_from_data=True)
            chart_ph.set_categories(cats)
            chart_ph.legend.position = 'b'
            chart_ph.height = 12
            chart_ph.width = 25
            ws_charts.add_chart(chart_ph, "E72")

            # === SECTION 5: PROFITS PAR JOUR ===
            ws_charts['A104'] = "💰 PROFITS PAR JOUR (OUT)"
            ws_charts['A104'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A106'] = "Jour"
            ws_charts['B106'] = "Profit (€)"
            row_ptr4 = 107
            profits_par_jour = aggs.get("profits_par_jour_out", pd.Series(dtype=float))
            jours_noms = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
            for d in range(7):
                ws_charts.cell(row=row_ptr4, column=1, value=jours_noms[d])
                ws_charts.cell(row=row_ptr4, column=2, value=float(round(profits_par_jour.get(d, 0.0), 2)))
                row_ptr4 += 1
            chart_pj = BarChart()
            chart_pj.title = "Profits par jour (OUT)"
            chart_pj.y_axis.title = "Profit (€)"
            chart_pj.x_axis.title = "Jour de la semaine"
            data = Reference(ws_charts, min_col=2, min_row=106, max_row=row_ptr4-1)
            cats = Reference(ws_charts, min_col=1, min_row=107, max_row=row_ptr4-1)
            chart_pj.add_data(data, titles_from_data=True)
            chart_pj.set_categories(cats)
            chart_pj.legend.position = 'b'
            chart_pj.height = 12
            chart_pj.width = 20
            ws_charts.add_chart(chart_pj, "E104")

            # === SECTION 6: PROFITS PAR MOIS ===
            ws_charts['A120'] = "💰 PROFITS PAR MOIS (OUT)"
            ws_charts['A120'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A122'] = "Mois"
            ws_charts['B122'] = "Profit (€)"
            row_ptr5 = 123
            profits_par_mois = aggs.get("profits_par_mois_out", pd.Series(dtype=float))
            mois_noms = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
            for m in range(1, 13):
                ws_charts.cell(row=row_ptr5, column=1, value=mois_noms[m-1])
                ws_charts.cell(row=row_ptr5, column=2, value=float(round(profits_par_mois.get(m, 0.0), 2)))
                row_ptr5 += 1
            chart_pm = BarChart()
            chart_pm.title = "Profits par mois (OUT)"
            chart_pm.y_axis.title = "Profit (€)"
            chart_pm.x_axis.title = "Mois de l'année"
            data = Reference(ws_charts, min_col=2, min_row=122, max_row=row_ptr5-1)
            cats = Reference(ws_charts, min_col=1, min_row=123, max_row=row_ptr5-1)
            chart_pm.add_data(data, titles_from_data=True)
            chart_pm.set_categories(cats)
            chart_pm.legend.position = 'b'
            chart_pm.height = 12
            chart_pm.width = 25
            ws_charts.add_chart(chart_pm, "E120")

            # === SECTION 7: TP/SL PAR HEURE ===
            ws_charts['A150'] = "🎯 NOMBRE DE TP/SL PAR HEURE"
            ws_charts['A150'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A152'] = "Heure"
            ws_charts['B152'] = "TP (nb)"
            ws_charts['C152'] = "SL (nb)"
            row_ptr6 = 153
            tp_h = aggs.get('tp_par_heure', pd.Series(dtype=int))
            sl_h = aggs.get('sl_par_heure', pd.Series(dtype=int))
            for h in range(24):
                ws_charts.cell(row=row_ptr6, column=1, value=h)
                ws_charts.cell(row=row_ptr6, column=2, value=int(tp_h.get(h, 0)))
                ws_charts.cell(row=row_ptr6, column=3, value=int(sl_h.get(h, 0)))
                row_ptr6 += 1
            chart_tpsl_h = BarChart()
            chart_tpsl_h.type = "col"
            chart_tpsl_h.grouping = "clustered"  # barres groupées TP vs SL
            chart_tpsl_h.title = "Nombre de TP/SL par heure (au dernier OUT)"
            chart_tpsl_h.y_axis.title = "Nombre"
            chart_tpsl_h.x_axis.title = "Heure (0-23)"
            data = Reference(ws_charts, min_col=2, min_row=152, max_col=3, max_row=row_ptr6-1)
            cats = Reference(ws_charts, min_col=1, min_row=153, max_row=row_ptr6-1)
            chart_tpsl_h.add_data(data, titles_from_data=True)
            chart_tpsl_h.set_categories(cats)
            chart_tpsl_h.legend.position = 'b'
            chart_tpsl_h.height = 12
            chart_tpsl_h.width = 25
            ws_charts.add_chart(chart_tpsl_h, "E150")

            # === SECTION 8: TP/SL PAR JOUR ===
            ws_charts['A185'] = "🎯 NOMBRE DE TP/SL PAR JOUR"
            ws_charts['A185'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A187'] = "Jour"
            ws_charts['B187'] = "TP (nb)"
            ws_charts['C187'] = "SL (nb)"
            row_ptr7 = 188
            tp_d = aggs.get('tp_par_jour', pd.Series(dtype=int))
            sl_d = aggs.get('sl_par_jour', pd.Series(dtype=int))
            for d in range(7):
                ws_charts.cell(row=row_ptr7, column=1, value=jours_noms[d])
                ws_charts.cell(row=row_ptr7, column=2, value=int(tp_d.get(d, 0)))
                ws_charts.cell(row=row_ptr7, column=3, value=int(sl_d.get(d, 0)))
                row_ptr7 += 1
            chart_tpsl_d = BarChart()
            chart_tpsl_d.type = "col"
            chart_tpsl_d.grouping = "clustered"
            chart_tpsl_d.title = "Nombre de TP/SL par jour"
            chart_tpsl_d.y_axis.title = "Nombre"
            chart_tpsl_d.x_axis.title = "Jour de la semaine"
            data = Reference(ws_charts, min_col=2, min_row=187, max_col=3, max_row=row_ptr7-1)
            cats = Reference(ws_charts, min_col=1, min_row=188, max_row=row_ptr7-1)
            chart_tpsl_d.add_data(data, titles_from_data=True)
            chart_tpsl_d.set_categories(cats)
            chart_tpsl_d.legend.position = 'b'
            chart_tpsl_d.height = 12
            chart_tpsl_d.width = 20
            ws_charts.add_chart(chart_tpsl_d, "E185")

            # === SECTION 9: TP/SL PAR MOIS ===
            ws_charts['A200'] = "🎯 NOMBRE DE TP/SL PAR MOIS"
            ws_charts['A200'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A202'] = "Mois"
            ws_charts['B202'] = "TP (nb)"
            ws_charts['C202'] = "SL (nb)"
            row_ptr8 = 203
            tp_m = aggs.get('tp_par_mois', pd.Series(dtype=int))
            sl_m = aggs.get('sl_par_mois', pd.Series(dtype=int))
            for m in range(1, 13):
                ws_charts.cell(row=row_ptr8, column=1, value=mois_noms[m-1])
                ws_charts.cell(row=row_ptr8, column=2, value=int(tp_m.get(m, 0)))
                ws_charts.cell(row=row_ptr8, column=3, value=int(sl_m.get(m, 0)))
                row_ptr8 += 1
            chart_tpsl_m = BarChart()
            chart_tpsl_m.type = "col"
            chart_tpsl_m.grouping = "clustered"
            chart_tpsl_m.title = "Nombre de TP/SL par mois"
            chart_tpsl_m.y_axis.title = "Nombre"
            chart_tpsl_m.x_axis.title = "Mois de l'année"
            data = Reference(ws_charts, min_col=2, min_row=202, max_col=3, max_row=row_ptr8-1)
            cats = Reference(ws_charts, min_col=1, min_row=203, max_row=row_ptr8-1)
            chart_tpsl_m.add_data(data, titles_from_data=True)
            chart_tpsl_m.set_categories(cats)
            chart_tpsl_m.legend.position = 'b'
            chart_tpsl_m.height = 12
            chart_tpsl_m.width = 25
            ws_charts.add_chart(chart_tpsl_m, "E200")

            # === SECTION 11: PROFITS VS PERTES GROUPÉS ===
            # Ancrage dynamique pour éviter tout chevauchement avec les sections précédentes
            base_anchor = max(ws_charts.max_row + 5, 260)
            # Heures
            start = base_anchor
            ws_charts[f'A{start}'] = "💰 PROFITS VS PERTES PAR HEURE (OUT)"
            ws_charts[f'A{start}'].font = Font(bold=True, color="366092", size=14)
            ws_charts[f'A{start+2}'] = "Heure"
            ws_charts[f'B{start+2}'] = "Profits (≥0)"
            ws_charts[f'C{start+2}'] = "Pertes (≤0)"
            rowp = start + 3
            profits_pos_h = aggs.get("profits_pos_par_heure_out")
            pertes_h = aggs.get("pertes_abs_par_heure_out")
            for h in range(24):
                ws_charts.cell(row=rowp, column=1, value=h)
                ws_charts.cell(row=rowp, column=2, value=float(round(float(profits_pos_h.get(h, 0.0)), 2)))
                ws_charts.cell(row=rowp, column=3, value=float(round(float(pertes_h.get(h, 0.0)), 2)))
                rowp += 1
            chart_ph_group = BarChart()
            chart_ph_group.type = "col"
            chart_ph_group.grouping = "clustered"
            chart_ph_group.title = "Profits vs Pertes par heure (OUT)"
            chart_ph_group.y_axis.title = "Montant (€)"
            chart_ph_group.x_axis.title = "Heure (0-23)"
            data = Reference(ws_charts, min_col=2, min_row=start+2, max_col=3, max_row=rowp-1)
            cats = Reference(ws_charts, min_col=1, min_row=start+3, max_row=rowp-1)
            chart_ph_group.add_data(data, titles_from_data=True)
            chart_ph_group.set_categories(cats)
            chart_ph_group.legend.position = 'b'
            chart_ph_group.height = 12
            chart_ph_group.width = 25
            ws_charts.add_chart(chart_ph_group, f"E{start}")

            # Jours
            start2 = max(ws_charts.max_row + 5, rowp + 15)
            ws_charts[f'A{start2}'] = "💰 PROFITS VS PERTES PAR JOUR (OUT)"
            ws_charts[f'A{start2}'].font = Font(bold=True, color="366092", size=14)
            ws_charts[f'A{start2+2}'] = "Jour"
            ws_charts[f'B{start2+2}'] = "Profits (≥0)"
            ws_charts[f'C{start2+2}'] = "Pertes (≤0)"
            rowp2 = start2 + 3
            profits_pos_d = aggs.get("profits_pos_par_jour_out")
            pertes_d = aggs.get("pertes_abs_par_jour_out")
            jours_noms = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
            for d in range(7):
                ws_charts.cell(row=rowp2, column=1, value=jours_noms[d])
                ws_charts.cell(row=rowp2, column=2, value=float(round(float(profits_pos_d.get(d, 0.0)), 2)))
                ws_charts.cell(row=rowp2, column=3, value=float(round(float(pertes_d.get(d, 0.0)), 2)))
                rowp2 += 1
            chart_pd_group = BarChart()
            chart_pd_group.type = "col"
            chart_pd_group.grouping = "clustered"
            chart_pd_group.title = "Profits vs Pertes par jour (OUT)"
            chart_pd_group.y_axis.title = "Montant (€)"
            chart_pd_group.x_axis.title = "Jour de la semaine"
            data = Reference(ws_charts, min_col=2, min_row=start2+2, max_col=3, max_row=rowp2-1)
            cats = Reference(ws_charts, min_col=1, min_row=start2+3, max_row=rowp2-1)
            chart_pd_group.add_data(data, titles_from_data=True)
            chart_pd_group.set_categories(cats)
            chart_pd_group.legend.position = 'b'
            chart_pd_group.height = 12
            chart_pd_group.width = 25
            ws_charts.add_chart(chart_pd_group, f"E{start2}")

            # Mois
            start3 = max(ws_charts.max_row + 5, rowp2 + 15)
            ws_charts[f'A{start3}'] = "💰 PROFITS VS PERTES PAR MOIS (OUT)"
            ws_charts[f'A{start3}'].font = Font(bold=True, color="366092", size=14)
            ws_charts[f'A{start3+2}'] = "Mois"
            ws_charts[f'B{start3+2}'] = "Profits (≥0)"
            ws_charts[f'C{start3+2}'] = "Pertes (≤0)"
            rowp3 = start3 + 3
            profits_pos_m = aggs.get("profits_pos_par_mois_out")
            pertes_m = aggs.get("pertes_abs_par_mois_out")
            mois_noms = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
            for m in range(1, 13):
                ws_charts.cell(row=rowp3, column=1, value=mois_noms[m-1])
                ws_charts.cell(row=rowp3, column=2, value=float(round(float(profits_pos_m.get(m, 0.0)), 2)))
                ws_charts.cell(row=rowp3, column=3, value=float(round(float(pertes_m.get(m, 0.0)), 2)))
                rowp3 += 1
            chart_pm_group = BarChart()
            chart_pm_group.type = "col"
            chart_pm_group.grouping = "clustered"
            chart_pm_group.title = "Profits vs Pertes par mois (OUT)"
            chart_pm_group.y_axis.title = "Montant (€)"
            chart_pm_group.x_axis.title = "Mois de l'année"
            data = Reference(ws_charts, min_col=2, min_row=start3+2, max_col=3, max_row=rowp3-1)
            cats = Reference(ws_charts, min_col=1, min_row=start3+3, max_row=rowp3-1)
            chart_pm_group.add_data(data, titles_from_data=True)
            chart_pm_group.set_categories(cats)
            chart_pm_group.legend.position = 'b'
            chart_pm_group.height = 12
            chart_pm_group.width = 25
            ws_charts.add_chart(chart_pm_group, f"E{start3}")

            # === SECTION 10: ÉVOLUTION DE LA SOMME CUMULÉE ===
            ws_charts['A220'] = "📈 ÉVOLUTION DE LA SOMME CUMULÉE"
            ws_charts['A220'].font = Font(bold=True, color="366092", size=14)
            ws_charts['A222'] = "Date"
            ws_charts['B222'] = "Solde cumulé (€)"
            
            # Récupérer les données de la colonne "solde_cumule" de df_final
            row_ptr9 = 223
            for _, row in df_final.iterrows():
                # Vérifier que la colonne datetime existe
                if 'datetime' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=1, value=row['datetime'])
                elif 'Date' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=1, value=row['Date'])
                elif 'date' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=1, value=row['date'])
                else:
                    print(f"[WARNING] Aucune colonne de date trouvée. Colonnes disponibles: {list(df_final.columns)}")
                    ws_charts.cell(row=row_ptr9, column=1, value=f"Ligne {row_ptr9-222}")
                
                # Vérifier que la colonne solde_cumule existe
                if 'solde_cumule' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=2, value=float(round(row['solde_cumule'], 2)))
                elif 'Solde_cumule' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=2, value=float(round(row['Solde_cumule'], 2)))
                elif 'solde_cumulé' in df_final.columns:
                    ws_charts.cell(row=row_ptr9, column=2, value=float(round(row['solde_cumulé'], 2)))
                else:
                    print(f"[WARNING] Colonne 'solde_cumule' non trouvée. Colonnes disponibles: {list(df_final.columns)}")
                    ws_charts.cell(row=row_ptr9, column=2, value=0.0)
                row_ptr9 += 1
            
            chart_cumul = LineChart()
            chart_cumul.title = "Évolution de la somme cumulée"
            chart_cumul.y_axis.title = "Solde cumulé (€)"
            chart_cumul.x_axis.title = "Date"
            data = Reference(ws_charts, min_col=2, min_row=222, max_row=row_ptr9-1)
            cats = Reference(ws_charts, min_col=1, min_row=223, max_row=row_ptr9-1)
            chart_cumul.add_data(data, titles_from_data=True)
            chart_cumul.set_categories(cats)
            chart_cumul.legend.position = 'b'
            chart_cumul.height = 15
            chart_cumul.width = 30
            ws_charts.add_chart(chart_cumul, "E220")
            
            # === ONGLET 2: DONNÉES BRUTES COMPLÈTES ===
            ws_data = wb.create_sheet("📋 Données Complètes")
            
            # Adapter les noms de colonnes selon le contenu
            df_final_copy = df_final.copy()
            colonnes_adaptees = {}
            
            # Vérifier s'il y a du Forex et des autres instruments
            if "Symbole_ordre" in df_final.columns:
                types_instruments = set()
                for symbole in df_final["Symbole_ordre"].unique():
                    type_inst = self.detecter_type_instrument(symbole)
                    types_instruments.add(type_inst)
                
                # Si on a seulement du Forex -> "Profit_pips"
                # Si on a seulement des autres -> "Profit_points"  
                # Si on a les deux -> "Profit_pips_points"
                if len(types_instruments) == 1:
                    type_unique = list(types_instruments)[0]
                    if type_unique == InstrumentType.FOREX:
                        colonnes_adaptees["Profit_pips"] = "Profit_pips"
                    else:
                        colonnes_adaptees["Profit_pips"] = "Profit_points"
                else:
                    # Mélange de types
                    colonnes_adaptees["Profit_pips"] = "Profit_pips_points"
            
            # Renommer les colonnes si nécessaire
            for ancienne, nouvelle in colonnes_adaptees.items():
                if ancienne in df_final_copy.columns:
                    df_final_copy = df_final_copy.rename(columns={ancienne: nouvelle})
            
            # Insérer toutes les données
            for r in dataframe_to_rows(df_final_copy, index=False, header=True):
                ws_data.append(r)
            
            # Formatage des en-têtes
            for cell in ws_data[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            print(f"[DEBUG] Data sheet created with {len(df_final)} rows")

            # === SURBRILLANCE: IN espacés de moins de 2 minutes (toutes les lignes du burst) ===
            try:
                if "Direction" in df_final.columns and "Heure d'ouverture" in df_final.columns:
                    df_flag = df_final[["Direction", "Heure d'ouverture"]].copy()
                    df_flag["__dt"] = pd.to_datetime(df_flag["Heure d'ouverture"], errors='coerce')
                    # Travailler uniquement sur les IN ordonnés par date/heure
                    df_in_only = df_flag[df_flag["Direction"].astype(str).str.lower() == "in"].copy()
                    df_in_only = df_in_only.sort_values("__dt")
                    # Diff vs précédent IN
                    diffs = df_in_only["__dt"].diff()
                    within_2min_curr = diffs <= pd.Timedelta(minutes=2)
                    # Diff vs suivant IN (pour marquer le précédent aussi)
                    diffs_next = df_in_only["__dt"].diff(periods=-1).abs()
                    within_2min_prev = diffs_next <= pd.Timedelta(minutes=2)
                    # Indices originaux en burst
                    indices_burst = set(df_in_only.index[within_2min_curr.fillna(False)].tolist() +
                                         df_in_only.index[within_2min_prev.fillna(False)].tolist())
                    if indices_burst:
                        # Construire un masque aligné sur l'ordre d'écriture (df_final_copy conserve l'ordre)
                        mask_by_index = df_final.index.to_series().isin(indices_burst)
                        # Appliquer un fond jaune sur chaque ligne concernée (ligne 1 = en-têtes)
                        fill_yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                        num_rows = len(df_final_copy)
                        num_cols = len(df_final_copy.columns)
                        for i in range(num_rows):  # i = 0 correspond à la première ligne de données (row 2 dans Excel)
                            if bool(mask_by_index.iloc[i]):
                                excel_row = i + 2
                                for j in range(1, num_cols + 1):
                                    ws_data.cell(row=excel_row, column=j).fill = fill_yellow
                        print(f"[DEBUG] Highlighted {sum(mask_by_index)} IN rows within 2 minutes bursts")
            except Exception as e:
                print(f"[WARNING] Failed to apply 2-min IN highlighting: {str(e)}")
            
            # === ONGLET 3: ANALYSE PAR INSTRUMENT ===
            if "Symbole_ordre" in df_final.columns:
                ws_instruments = wb.create_sheet("📈 Analyse par Instrument")
                
                # Analyser les performances par instrument
                # Compter les trades complets par instrument (clés uniques)
                trades_par_instrument = df_final.groupby("Symbole_ordre")["Cle_Match"].nunique()
                
                # Calculer les profits par instrument
                profits_par_instrument = df_final.groupby("Symbole_ordre").agg({
                    'Profit': ['sum', 'mean'],
                    'Profit_pips': ['sum', 'mean']
                }).round(2)
                
                # Combiner les résultats
                analyse_instruments = pd.DataFrame({
                    'Nb_Trades': trades_par_instrument,
                    'Profit_Total': profits_par_instrument[('Profit', 'sum')],
                    'Profit_Moyen': profits_par_instrument[('Profit', 'mean')],
                    'Pips_Total': profits_par_instrument[('Profit_pips', 'sum')],
                    'Pips_Moyen': profits_par_instrument[('Profit_pips', 'mean')]
                }).reset_index()
                
                analyse_instruments = analyse_instruments.sort_values('Profit_Total', ascending=False)
                
                # Ajouter le type d'instrument
                analyse_instruments['Type_Instrument'] = analyse_instruments['Symbole_ordre'].apply(self.detecter_type_instrument)
                
                # En-têtes
                headers_instruments = ['Instrument', 'Type', 'Nb Trades', 'Profit Total (€)', 'Profit Moyen (€)', 'Pips/Points Total', 'Pips/Points Moyen']
                for col_idx, header in enumerate(headers_instruments, 1):
                    cell = ws_instruments.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Données
                for row_idx, (_, row) in enumerate(analyse_instruments.iterrows(), 2):
                    ws_instruments.cell(row=row_idx, column=1, value=row['Symbole_ordre'])
                    ws_instruments.cell(row=row_idx, column=2, value=str(row['Type_Instrument'].value).upper())
                    ws_instruments.cell(row=row_idx, column=3, value=int(row['Nb_Trades']))
                    ws_instruments.cell(row=row_idx, column=4, value=float(row['Profit_Total']))
                    ws_instruments.cell(row=row_idx, column=5, value=float(row['Profit_Moyen']))
                    ws_instruments.cell(row=row_idx, column=6, value=float(row['Pips_Total']))
                    ws_instruments.cell(row=row_idx, column=7, value=float(row['Pips_Moyen']))
                
                print(f"[DEBUG] Instruments analysis sheet created")
            
            # === ONGLET 4: ANALYSE PAR TYPE D'INSTRUMENT ===
            if "Symbole_ordre" in df_final.columns:
                ws_types = wb.create_sheet("🏷️ Analyse par Type")
                
                # Ajouter une colonne temporaire pour le type d'instrument
                df_final_copy = df_final.copy()
                df_final_copy['Type_Instrument'] = df_final_copy['Symbole_ordre'].apply(self.detecter_type_instrument)
                
                # Convertir les enums en chaînes pour le groupby
                df_final_copy['Type_Instrument_Str'] = df_final_copy['Type_Instrument'].apply(lambda x: x.value)
                
                # Analyser par type d'instrument
                # Compter les trades complets par type (clés uniques)
                trades_par_type = df_final_copy.groupby("Type_Instrument_Str")["Cle_Match"].nunique()
                
                # Calculer les profits par type
                profits_par_type = df_final_copy.groupby("Type_Instrument_Str").agg({
                    'Profit': ['sum', 'mean'],
                    'Profit_pips': ['sum', 'mean']
                }).round(2)
                
                # Combiner les résultats
                analyse_types = pd.DataFrame({
                    'Nb_Trades': trades_par_type,
                    'Profit_Total': profits_par_type[('Profit', 'sum')],
                    'Profit_Moyen': profits_par_type[('Profit', 'mean')],
                    'Pips_Total': profits_par_type[('Profit_pips', 'sum')],
                    'Pips_Moyen': profits_par_type[('Profit_pips', 'mean')]
                }).reset_index()
                
                analyse_types = analyse_types.sort_values('Profit_Total', ascending=False)
                
                # En-têtes
                headers_types = ['Type d\'Instrument', 'Nb Trades', 'Profit Total (€)', 'Profit Moyen (€)', 'Pips/Points Total', 'Pips/Points Moyen']
                for col_idx, header in enumerate(headers_types, 1):
                    cell = ws_types.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Données avec icônes selon le type
                type_icons = {
                    InstrumentType.FOREX: '💱',
                    InstrumentType.METAUX: '🥇',
                    InstrumentType.INDICES: '📊',
                    InstrumentType.CRYPTO: '₿',
                    InstrumentType.ENERGIE: '🛢️',
                    InstrumentType.ACTIONS: '📈'
                }
                
                for row_idx, (_, row) in enumerate(analyse_types.iterrows(), 2):
                    type_inst_str = str(row['Type_Instrument_Str'])  # Prendre la vraie valeur du type
                    # Trouver l'enum correspondant pour l'icône
                    type_inst_enum = None
                    for enum_val in InstrumentType:
                        if enum_val.value == type_inst_str:
                            type_inst_enum = enum_val
                            break
                    
                    icon = type_icons.get(type_inst_enum, '📈')
                    
                    # Afficher le nom complet avec l'icône
                    nom_complet = {
                        'forex': 'FOREX',
                        'metaux': 'MÉTAUX', 
                        'indices': 'INDICES',
                        'crypto': 'CRYPTO',
                        'energie': 'ÉNERGIE',
                        'actions': 'ACTIONS'
                    }.get(type_inst_str, type_inst_str.upper())
                    
                    ws_types.cell(row=row_idx, column=1, value=f"{icon} {nom_complet}")
                    ws_types.cell(row=row_idx, column=2, value=int(row['Nb_Trades']))
                    ws_types.cell(row=row_idx, column=3, value=float(row['Profit_Total']))
                    ws_types.cell(row=row_idx, column=4, value=float(row['Profit_Moyen']))
                    ws_types.cell(row=row_idx, column=5, value=float(row['Pips_Total']))
                    ws_types.cell(row=row_idx, column=6, value=float(row['Pips_Moyen']))
                
                print(f"[DEBUG] Instrument types analysis sheet created")
            
            # === ONGLET 5: DÉTAIL PAR INSTRUMENT ===
            if "Symbole_ordre" in df_final.columns:
                # Obtenir la liste unique des instruments
                instruments_uniques = df_final["Symbole_ordre"].unique()
                
                for instrument in instruments_uniques:
                    # Créer un nom d'onglet sécurisé (Excel limite à 31 caractères)
                    nom_onglet = f"📊 {instrument[:25]}" if len(instrument) > 25 else f"📊 {instrument}"
                    
                    # Éviter les doublons d'onglets
                    if nom_onglet in [ws.title for ws in wb.worksheets]:
                        nom_onglet = f"📊 {instrument[:20]}_{hash(instrument) % 1000}"
                    
                    try:
                        ws_instrument = wb.create_sheet(nom_onglet)
                        
                        # Filtrer les données pour cet instrument
                        df_instrument = df_final[df_final["Symbole_ordre"] == instrument].copy()
                        
                        # Titre de l'instrument
                        ws_instrument.merge_cells('A1:H1')
                        cell_titre = ws_instrument['A1']
                        cell_titre.value = f"ANALYSE DÉTAILLÉE - {instrument.upper()}"
                        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
                        cell_titre.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell_titre.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Déterminer le type d'instrument pour l'affichage
                        type_instrument = self.detecter_type_instrument(instrument)
                        is_forex = (type_instrument == InstrumentType.FOREX)
                        unite_mesure = "Pips" if is_forex else "Points"
                        
                        # Statistiques de l'instrument (basées sur les trades complets)
                        nb_trades_complets = df_instrument["Cle_Match"].nunique()
                        nb_operations = len(df_instrument)
                        
                        # Calculer les trades gagnants/perdants basés sur les trades complets
                        trades_complets_instrument = df_instrument.groupby("Cle_Match").agg({
                            "Profit": "sum"
                        }).reset_index()
                        
                        trades_gagnants = len(trades_complets_instrument[trades_complets_instrument["Profit"] > 0])
                        trades_perdants = len(trades_complets_instrument[trades_complets_instrument["Profit"] < 0])
                        profit_total = df_instrument["Profit"].sum()
                        pips_total = df_instrument["Profit_pips"].sum()
                        taux_reussite = (trades_gagnants / nb_trades_complets * 100) if nb_trades_complets > 0 else 0
                        
                        # Tableau des statistiques
                        stats_instrument = [
                            ["📊 STATISTIQUES DE L'INSTRUMENT", ""],
                            ["", ""],
                            ["📈 Nombre total de trades complets", nb_trades_complets],
                            ["📈 Nombre total d'opérations", nb_operations],
                            ["✅ Trades gagnants", trades_gagnants],
                            ["❌ Trades perdants", trades_perdants],
                            ["🎯 Taux de réussite", f"{taux_reussite:.1f} %"],
                            ["💰 Profit total", f"{profit_total:,.2f} €"],
                            [f"🎯 {unite_mesure} totaux", f"{pips_total:,.2f}"],
                            ["", ""],
                            ["📋 DÉTAIL DES TRADES", ""],
                            ["", ""]
                        ]
                        
                        for row_idx, (label, value) in enumerate(stats_instrument, 3):
                            ws_instrument[f'A{row_idx}'] = label
                            ws_instrument[f'B{row_idx}'] = value
                            
                            if any(word in label for word in ["STATISTIQUES", "DÉTAIL"]):
                                ws_instrument[f'A{row_idx}'].font = Font(bold=True, color="366092")
                                ws_instrument[f'A{row_idx}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        
                        # En-têtes des colonnes de données (adapter selon le type)
                        headers_data = list(df_instrument.columns)
                        headers_adaptes = []
                        for header in headers_data:
                            if header == "Profit_pips":
                                header_adapte = f"Profit_{unite_mesure.lower()}"
                            else:
                                header_adapte = header
                            headers_adaptes.append(header_adapte)
                        
                        for col_idx, header in enumerate(headers_adaptes, 1):
                            cell = ws_instrument.cell(row=len(stats_instrument) + 3, column=col_idx, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                        
                        # Données
                        for row_idx, (_, row) in enumerate(df_instrument.iterrows(), len(stats_instrument) + 4):
                            for col_idx, value in enumerate(row, 1):
                                ws_instrument.cell(row=row_idx, column=col_idx, value=value)
                        
                        print(f"[DEBUG] Created detailed sheet for {instrument}")

                        # Tableau sessions pour cet instrument
                        try:
                            bloc_pair = sessions.get("sessions_par_pair", {}).get(instrument, {}) if 'sessions' in locals() else {}
                            if bloc_pair:
                                start_row = ws_instrument.max_row + 2
                                ws_instrument[f'A{start_row}'] = "🌍 PERFORMANCE PAR SESSION"
                                ws_instrument[f'A{start_row}'].font = Font(bold=True, color="366092")
                                headers = ["Session", "IN (nb)", "Taux réussite IN (%)", "PnL OUT (€)", "TP (nb)", "SL (nb)"]
                                for i, h in enumerate(headers, 0):
                                    cell = ws_instrument.cell(row=start_row+2, column=1+i, value=h)
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center")
                                lignes = ["Asie","Europe","Amérique"]
                                for r_idx, sess in enumerate(lignes, start=start_row+3):
                                    ws_instrument.cell(row=r_idx, column=1, value=sess)
                                    ws_instrument.cell(row=r_idx, column=2, value=int(bloc_pair.get('in_count',{}).get(sess,0)))
                                    ws_instrument.cell(row=r_idx, column=3, value=float(bloc_pair.get('taux_reussite_in_pct',{}).get(sess,0)))
                                    ws_instrument.cell(row=r_idx, column=4, value=float(bloc_pair.get('pnl_out',{}).get(sess,0)))
                                    ws_instrument.cell(row=r_idx, column=5, value=int(bloc_pair.get('tp_out',{}).get(sess,0)))
                                    ws_instrument.cell(row=r_idx, column=6, value=int(bloc_pair.get('sl_out',{}).get(sess,0)))
                        except Exception as e:
                            print(f"[WARNING] Session table for instrument {instrument} failed: {e}")
                        
                        # Bloc Patterns pour cet instrument
                        try:
                            patterns_pair = self.calculer_patterns(df_instrument, n_permutations=500)
                            start_row_patterns = ws_instrument.max_row + 2
                            ws_instrument[f'A{start_row_patterns}'] = "🧩 PATTERNS (PAIR)"
                            ws_instrument[f'A{start_row_patterns}'].font = Font(bold=True, color="366092")
                            headers_p = ["Items", "Count", "Support", "Confidence", "Lift", "p-value", "Signif"]
                            for i, h in enumerate(headers_p, 0):
                                cell = ws_instrument.cell(row=start_row_patterns+2, column=1+i, value=h)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")
                            # Favorables
                            ws_instrument.cell(row=start_row_patterns+3, column=1, value="Favorables (⇒ TP)")
                            ws_instrument.cell(row=start_row_patterns+3, column=1).font = Font(bold=True)
                            rowp = start_row_patterns + 4
                            for r in patterns_pair.get('top_tp', [])[:10]:
                                ws_instrument.cell(row=rowp, column=1, value=r['items'])
                                ws_instrument.cell(row=rowp, column=2, value=r['count'])
                                ws_instrument.cell(row=rowp, column=3, value=r['support'])
                                ws_instrument.cell(row=rowp, column=4, value=r['confidence'])
                                ws_instrument.cell(row=rowp, column=5, value=r['lift'])
                                pval = r.get('p_value')
                                ws_instrument.cell(row=rowp, column=6, value=pval)
                                signif = "***" if pval is not None and pval <= 0.01 else "**" if pval is not None and pval <= 0.05 else "*" if pval is not None and pval <= 0.10 else "."
                                ws_instrument.cell(row=rowp, column=7, value=signif)
                                rowp += 1
                            # Défavorable
                            rowp += 1
                            ws_instrument.cell(row=rowp, column=1, value="Défavorables (⇒ SL)")
                            ws_instrument.cell(row=rowp, column=1).font = Font(bold=True)
                            rowp += 1
                            for r in patterns_pair.get('top_sl', [])[:10]:
                                ws_instrument.cell(row=rowp, column=1, value=r['items'])
                                ws_instrument.cell(row=rowp, column=2, value=r['count'])
                                ws_instrument.cell(row=rowp, column=3, value=r['support'])
                                ws_instrument.cell(row=rowp, column=4, value=r['confidence'])
                                ws_instrument.cell(row=rowp, column=5, value=r['lift'])
                                pval = r.get('p_value')
                                ws_instrument.cell(row=rowp, column=6, value=pval)
                                signif = "***" if pval is not None and pval <= 0.01 else "**" if pval is not None and pval <= 0.05 else "*" if pval is not None and pval <= 0.10 else "."
                                ws_instrument.cell(row=rowp, column=7, value=signif)
                                rowp += 1
                        except Exception as e:
                            print(f"[WARNING] Pair patterns failed for {instrument}: {e}")

                    except Exception as e:
                        print(f"[WARNING] Could not create sheet for {instrument}: {str(e)}")
                        continue

            # === ONGLET 6: PATTERNS (MVP étendu) ===
            try:
                patterns = self.calculer_patterns(df_final, n_permutations=500, max_itemset_size=3)
                ws_patterns = wb.create_sheet("🧩 Patterns")
                # Explications
                ws_patterns['A1'] = "🧩 Détection de patterns (règles d'association)"
                ws_patterns['A1'].font = Font(bold=True, color="366092", size=14)
                exp_lines = [
                    "Items utilisés: DIR (buy/sell), SESSION (Asie/Europe/Amérique), heure d'ouverture (H[..]) et durée (D..).",
                    "DIR=buy/sell: sens d'ouverture du trade.",
                    "SESSION=Asie/Europe/Amérique: session du marché à l'ouverture (H 0–7 / 8–15 / 16–23).",
                    "Heure (H[..]): plages horaires d'ouverture: H[0-7], H[8-11], H[12-15], H[16-19], H[20-23].",
                    "Durée (D..): temps entre IN et dernier OUT: D<30m, D30-120m, D2-6h, D6-12h, D12-24h, >24h.",
                    "Count: nombre de trades contenant le pattern (itemset).",
                    "Support: proportion de trades contenant le pattern = Count / N (N = total de trades complets).",
                    "Confidence: probabilité de TP (ou SL) sachant le pattern = Count(pattern ∪ TP) / Count(pattern).",
                    "Lift: surperformance relative = Confidence / P(TP) (ou / P(SL)). Lift > 1 => pattern informatif.",
                    "p-value (permutation): probabilité d'obtenir une confidence ≥ observée si la cible (TP/SL) était aléatoire. Plus c'est petit, plus le pattern est significatif.",
                    "Méthode de permutation (n=500): on mélange aléatoirement les étiquettes TP/SL entre trades, on recalcule la confidence à chaque mélange, la p-value est la part des mélanges ≥ à la confidence observée.",
                    "q-value (FDR Benjamini–Hochberg): p-value ajustée pour multiplicité. C'est la proportion d'hypothèses fausses attendue parmi les règles déclarées significatives.",
                    "Interprétation: p≤0.01 (***), p≤0.05 (**), p≤0.10 (*) sinon (.). Ces seuils sont indicatifs, à croiser avec support et lift.",
                    "But de cette section: proposer des contextes (items) où la probabilité de TP (ou de SL) diffère significativement du taux global.",
                ]
                for i, line in enumerate(exp_lines, start=2):
                    ws_patterns[f'A{i}'] = line
                # Titre bloc TP
                start_tp_row = len(exp_lines) + 3
                ws_patterns[f'A{start_tp_row}'] = "TOP PATTERNS FAVORABLES (⇒ TP) — triés par p-value"
                ws_patterns[f'A{start_tp_row}'].font = Font(bold=True, color="366092", size=14)
                headers = ["Items", "Count", "Support", "Confidence", "Lift", "p-value", "q-value", "Signif"]
                for i, h in enumerate(headers, 1):
                    cell = ws_patterns.cell(row=start_tp_row+2, column=i, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                row = start_tp_row + 3
                for r in patterns.get('top_tp', []):
                    ws_patterns.cell(row=row, column=1, value=r['items'])
                    ws_patterns.cell(row=row, column=2, value=r['count'])
                    ws_patterns.cell(row=row, column=3, value=r['support'])
                    ws_patterns.cell(row=row, column=4, value=r['confidence'])
                    ws_patterns.cell(row=row, column=5, value=r['lift'])
                    pval = r.get('p_value')
                    ws_patterns.cell(row=row, column=6, value=pval)
                    ws_patterns.cell(row=row, column=7, value=r.get('q_value'))
                    signif = "***" if pval is not None and pval <= 0.01 else "**" if pval is not None and pval <= 0.05 else "*" if pval is not None and pval <= 0.10 else "."
                    ws_patterns.cell(row=row, column=8, value=signif)
                    row += 1

                row += 2
                ws_patterns.cell(row=row, column=1, value="TOP PATTERNS DÉFAVORABLES (⇒ SL)")
                ws_patterns.cell(row=row, column=1).font = Font(bold=True, color="366092", size=14)
                row += 2
                for i, h in enumerate(headers, 1):
                    cell = ws_patterns.cell(row=row, column=i, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                row += 1
                for r in patterns.get('top_sl', []):
                    ws_patterns.cell(row=row, column=1, value=r['items'])
                    ws_patterns.cell(row=row, column=2, value=r['count'])
                    ws_patterns.cell(row=row, column=3, value=r['support'])
                    ws_patterns.cell(row=row, column=4, value=r['confidence'])
                    ws_patterns.cell(row=row, column=5, value=r['lift'])
                    pval = r.get('p_value')
                    ws_patterns.cell(row=row, column=6, value=pval)
                    ws_patterns.cell(row=row, column=7, value=r.get('q_value'))
                    signif = "***" if pval is not None and pval <= 0.01 else "**" if pval is not None and pval <= 0.05 else "*" if pval is not None and pval <= 0.10 else "."
                    ws_patterns.cell(row=row, column=8, value=signif)
                    row += 1

                # Bloc modèle d'influence (logit)
                try:
                    ws_patterns.cell(row=row+2, column=1, value="📐 Modèle d'influence (logistique TP vs SL)")
                    ws_patterns.cell(row=row+2, column=1).font = Font(bold=True, color="366092", size=14)
                    # Explications du modèle
                    explain = [
                        "Feature: variable explicative binaire issue du one-hot encoding (ex: DIR_buy, SESSION_Europe, H[8-11], D2-6h).",
                        "Interactions: produits de deux features (ex: DIR_buy*SESSION_Europe) pour capturer des effets combinés.",
                        "Coef: effet (log-odds). Positif → augmente les chances de TP; négatif → augmente les chances de SL.",
                        "Odds Ratio: exp(coef). >1 → favorable au TP; <1 → défavorable. Ex: 1.30 = +30% sur les odds de TP.",
                        "p-value: significativité statistique de l'effet (si disponible).",
                        "Utilité: identifier les paramètres les plus influents globalement, au-delà des règles ponctuelles.",
                        "",
                        "Dictionnaire des features (comment lire):",
                        "- DIR_buy / DIR_sell: sens du trade à l'ouverture (achat/vente).",
                        "- SESSION_Asie / SESSION_Europe / SESSION_Amérique: session de marché lors de l'ouverture.",
                        "- H[0-7], H[8-11], H[12-15], H[16-19], H[20-23]: plage horaire d'ouverture du trade (UTC ou heure des données).",
                        "- D<30m, D30-120m, D2-6h, D6-12h, D12-24h, >24h: durée entre IN et dernier OUT.",
                        "- A*B (ex: DIR_buy*H[8-11]): effet spécifique quand A et B sont vrais simultanément (interaction).",
                    ]
                    for i, line in enumerate(explain, start=1):
                        ws_patterns.cell(row=row+2+i, column=1, value=line)
                    infl = self.calculer_modele_influence(df_final)
                    if not infl.empty:
                        base_row = row + 2 + len(explain) + 2
                        headers_m = ["Feature", "Coef", "Odds Ratio", "p-value"]
                        for i, h in enumerate(headers_m, 1):
                            cell = ws_patterns.cell(row=base_row, column=i, value=h)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                        r0 = base_row + 1
                        for _, rr in infl.iterrows():
                            ws_patterns.cell(row=r0, column=1, value=str(rr['feature']))
                            ws_patterns.cell(row=r0, column=2, value=float(rr['coef']))
                            ws_patterns.cell(row=r0, column=3, value=float(rr['odds_ratio']))
                            # p_value peut être None dans certains backends
                            pv = rr.get('p_value')
                            ws_patterns.cell(row=r0, column=4, value=(float(pv) if pv is not None else None))
                            r0 += 1
                except Exception as e:
                    print(f"[WARNING] Influence model block failed: {e}")
            except Exception as e:
                print(f"[WARNING] Patterns sheet creation failed: {e}")
            
            # Tableau: PERFORMANCE PAR SESSION (TOTAL) en bas du Résumé
            try:
                bloc_total = sessions.get("sessions_total", {}) if 'sessions' in locals() else {}
                if bloc_total:
                    last_row = ws_resume.max_row + 2
                    ws_resume[f'A{last_row}'] = "🌍 PERFORMANCE PAR SESSION (TOTAL)"
                    ws_resume[f'A{last_row}'].font = Font(bold=True, color="366092")
                    headers = ["Session", "IN (nb)", "Taux réussite IN (%)", "PnL OUT (€)", "TP (nb)", "SL (nb)"]
                    for i, h in enumerate(headers, 0):
                        cell = ws_resume.cell(row=last_row+2, column=1+i, value=h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    lignes = ["Asie","Europe","Amérique"]
                    for r_idx, sess in enumerate(lignes, start=last_row+3):
                        ws_resume.cell(row=r_idx, column=1, value=sess)
                        ws_resume.cell(row=r_idx, column=2, value=int(bloc_total.get('in_count',{}).get(sess,0)))
                        ws_resume.cell(row=r_idx, column=3, value=float(bloc_total.get('taux_reussite_in_pct',{}).get(sess,0)))
                        ws_resume.cell(row=r_idx, column=4, value=float(bloc_total.get('pnl_out',{}).get(sess,0)))
                        ws_resume.cell(row=r_idx, column=5, value=int(bloc_total.get('tp_out',{}).get(sess,0)))
                        ws_resume.cell(row=r_idx, column=6, value=int(bloc_total.get('sl_out',{}).get(sess,0)))
            except Exception as e:
                print(f"[WARNING] Session total table creation failed: {e}")

            # Ajuster la largeur des colonnes
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            suffix = f"_{filter_type.upper()}" if filter_type else "_UNIFIED"
            fichier_rapport = os.path.join(reports_folder, f"RAPPORT_TRADING{suffix}_{timestamp}.xlsx")
            wb.save(fichier_rapport)
            
            print(f"[DEBUG] Excel report saved successfully: {fichier_rapport}")
            return fichier_rapport
            
        except Exception as e:
            print(f"[ERROR] Error creating Excel report: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors de la création du rapport Excel: {str(e)}")

def main():
    """Fonction principale pour tester le script"""
    analyzer = TradingAnalyzer(solde_initial=10000)
    
    # Exemple d'utilisation
    print("=== ANALYSEUR DE TRADING UNIFIÉ ===")
    print("1. Analyse Forex uniquement")
    print("2. Analyse autres instruments uniquement")
    print("3. Analyse complète (tous instruments)")
    print("4. Quitter")
    
    choix = input("\nChoisissez une option (1-4): ")
    
    if choix == "1":
        print("Analyse Forex sélectionnée")
        # Ici vous ajouteriez la logique pour sélectionner les fichiers
    elif choix == "2":
        print("Analyse autres instruments sélectionnée")
    elif choix == "3":
        print("Analyse complète sélectionnée")
    elif choix == "4":
        print("Au revoir!")
    else:
        print("Option invalide")

if __name__ == "__main__":
    main()