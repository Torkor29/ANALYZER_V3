import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, PieChart, BarChart
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum
import hashlib

class InstrumentType(Enum):
    """Types d'instruments financiers"""
    FOREX = "forex"
    METAUX = "metaux"
    INDICES = "indices"
    CRYPTO = "crypto"
    ENERGIE = "energie"
    ACTIONS = "actions"

class TradingAnalyzerImproved:
    """
    Analyseur unifié amélioré avec clé de jointure renforcée
    et feuilles spécialisées par type d'instrument
    """
    
    def __init__(self, solde_initial=10000):
        self.solde_initial = solde_initial
        self.statistiques_fichiers = {}
        
        # Configuration des instruments
        self.forex_pairs = [
            "eurusd", "gbpusd", "usdchf", "usdjpy", "usdcad", "audusd", "nzdusd",
            "eurjpy", "gbpjpy", "audjpy", "cadjpy", "chfjpy", "nzdjpy",
            "eurgbp", "euraud", "eurcad", "eurchf", "eurnzd",
            "gbpaud", "gbpcad", "gbpchf", "gbpnzd",
            "audcad", "audchf", "audnzd",
            "cadchf", "nzdcad", "nzdchf"
        ]
        
        self.instrument_patterns = {
            InstrumentType.METAUX: ["gold", "xauusd", "xau", "or", "silver", "xagusd", "xag", "argent", "platinum", "xptusd", "palladium", "xpdusd"],
            InstrumentType.INDICES: ["dax", "cac", "sp500", "dow", "nasdaq", "ftse", "nikkei", "asx", "us30", "us500", "ger30", "fra40", "uk100"],
            InstrumentType.CRYPTO: ["btc", "eth", "ltc", "xrp", "ada", "dot", "bitcoin", "ethereum", "crypto"],
            InstrumentType.ENERGIE: ["oil", "wti", "brent", "petrol", "crude", "gas", "natural"]
        }
    
    def process_files(self, file_paths, task_id, task_status, instrument_filter=None):
        """Traite une liste de fichiers Excel pour l'analyse"""
        try:
            print(f"[DEBUG] Starting improved analysis with {len(file_paths)} files")
            print(f"[DEBUG] Instrument filter: {instrument_filter}")
            
            tous_les_resultats = []
            total_files = len(file_paths)
            
            for i, file_path in enumerate(file_paths):
                progress = 20 + (i / total_files) * 40
                task_status[task_id]['progress'] = int(progress)
                task_status[task_id]['message'] = f'Traitement du fichier {i+1}/{total_files}...'
                
                print(f"[DEBUG] Processing file {i+1}/{total_files}: {os.path.basename(file_path)}")
                
                df_result, erreur, exclus, doublons = self.process_single_file(
                    file_path, i+1, total_files, instrument_filter
                )
                
                if df_result is not None and len(df_result) > 0:
                    tous_les_resultats.append(df_result)
                    
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': len(df_result),
                        'exclus': exclus,
                        'doublons': doublons,
                        'erreur': erreur
                    }
                    print(f"[DEBUG] File processed successfully: {len(df_result)} trades, {exclus} excluded")
                else:
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': 0,
                        'exclus': 0,
                        'doublons': 0,
                        'erreur': erreur or "Aucune donnée trouvée"
                    }
                    print(f"[DEBUG] File failed: {erreur}")
            
            if not tous_les_resultats:
                print(f"[DEBUG] No valid data found in any file")
                return None
            
            task_status[task_id]['progress'] = 60
            task_status[task_id]['message'] = 'Fusion des données et calculs des intérêts composés...'
            
            print(f"[DEBUG] Starting fusion and compound interest calculations")
            df_final = self.fusionner_et_calculer_cumuls(tous_les_resultats)
            print(f"[DEBUG] Fusion completed: {len(df_final)} total trades")
            
            task_status[task_id]['progress'] = 75
            task_status[task_id]['message'] = 'Calculs des statistiques avancées...'
            
            return df_final
            
        except Exception as e:
            print(f"[ERROR] Error in process_files: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors du traitement des fichiers: {str(e)}")
    
    def process_single_file(self, file_path, file_number, total_files, instrument_filter=None):
        """Traite un seul fichier Excel avec clé de jointure renforcée"""
        try:
            print(f"[DEBUG] Starting to process file: {file_path}")
            df = pd.read_excel(file_path, sheet_name=0, header=None)
            print(f"[DEBUG] File read successfully, shape: {df.shape}")
            
            ligne_ordres = self.trouver_ligne(df, "ordre")
            ligne_transactions = self.trouver_ligne(df, "transaction")
            print(f"[DEBUG] Found ordre line at: {ligne_ordres}, transaction line at: {ligne_transactions}")

            header_ordres = df.iloc[ligne_ordres + 1]
            ordres_df = df.iloc[ligne_ordres + 2 : ligne_transactions].copy()
            ordres_df.columns = header_ordres
            ordres_df.reset_index(drop=True, inplace=True)

            header_transactions = df.iloc[ligne_transactions + 1]
            transactions_df = df.iloc[ligne_transactions + 2 :].copy()
            transactions_df.columns = header_transactions
            transactions_df = transactions_df[transactions_df.iloc[:, 0].notna()]
            transactions_df.reset_index(drop=True, inplace=True)
            
            print(f"[DEBUG] Ordres shape: {ordres_df.shape}, Transactions shape: {transactions_df.shape}")

            if len(ordres_df.columns) < 2 or len(transactions_df.columns) < 2:
                return None, "Pas assez de colonnes dans les données", 0, 0

            # CLÉ DE JOINTURE RENFORCÉE
            ordres_df["__clé_renforcée__"] = self.creer_cle_jointure_renforcee(ordres_df)
            transactions_df["__clé_renforcée__"] = self.creer_cle_jointure_renforcee(transactions_df)
            
            if "Prix" in transactions_df.columns:
                transactions_df.rename(columns={"Prix": "Prix_transaction"}, inplace=True)

            fusion_df = pd.merge(ordres_df, transactions_df, on="__clé_renforcée__", suffixes=('_ordre', '_transaction'))
            print(f"[DEBUG] Merged dataframe shape: {fusion_df.shape}")
            
            avant_filtrage = len(fusion_df)

            if "Symbole_ordre" in fusion_df.columns:
                print(f"[DEBUG] Applying instrument filter: {instrument_filter}")
                fusion_df = self.filtrer_par_instrument(fusion_df, instrument_filter)
                apres_filtrage = len(fusion_df)
                print(f"[DEBUG] After filtering: {apres_filtrage} rows (excluded: {avant_filtrage - apres_filtrage})")
                
                if len(fusion_df) == 0:
                    return None, f"Aucun instrument {instrument_filter} trouvé", avant_filtrage - apres_filtrage, 0

            # Conversions des colonnes numériques
            print(f"[DEBUG] Converting numeric columns...")
            fusion_df["Profit"] = self.safe_convert_to_float(fusion_df["Profit"])
            fusion_df["Prix_transaction"] = self.safe_convert_to_float(fusion_df["Prix_transaction"])
            
            if "T / P" in fusion_df.columns:
                fusion_df["T / P"] = self.safe_convert_to_float(fusion_df["T / P"])
            if "S / L" in fusion_df.columns:
                fusion_df["S / L"] = self.safe_convert_to_float(fusion_df["S / L"])

            fusion_df["Volume_ordre"] = fusion_df["Volume_ordre"].astype(str)
            fusion_df["Symbole_ordre"] = fusion_df["Symbole_ordre"].astype(str)
            fusion_df["Cle_Match"] = None

            # Logique de matching renforcée
            print(f"[DEBUG] Applying enhanced matching logic...")
            self.apply_matching_logic_renforcee(fusion_df)
            
            # Calculs spécialisés par type d'instrument
            print(f"[DEBUG] Calculating specialized metrics...")
            fusion_df = self.calculer_metriques_specialisees(fusion_df)
            
            # Nettoyage et sélection des colonnes finales (MÉTRIQUES UNIFIÉES)
            colonnes_a_garder = [
                "Heure d'ouverture", "Ordre_ordre", "Symbole_ordre", "Type_ordre", 
                "Volume_ordre", "S / L", "T / P", "Direction", "Prix_transaction",
                "Profit", "Cle_Match", "Type_Instrument", "Profit_compose", 
                "Profit_cumule", "Solde_cumule", "Drawdown_pct", "Drawdown_euros"
            ]
            
            colonnes_finales = [col for col in colonnes_a_garder if col in fusion_df.columns]
            fusion_df = fusion_df[colonnes_finales]
            
            # Suppression des doublons
            avant_dedoublonnage = len(fusion_df)
            fusion_df = fusion_df.drop_duplicates().reset_index(drop=True)
            apres_dedoublonnage = len(fusion_df)
            doublons_supprimes = avant_dedoublonnage - apres_dedoublonnage
            
            print(f"[DEBUG] File processing completed: {len(fusion_df)} final trades")
            
            return fusion_df, "Succès", avant_filtrage - apres_filtrage, doublons_supprimes
            
        except Exception as e:
            print(f"[ERROR] Error processing file {file_path}: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            return None, str(e), 0, 0
    
    def creer_cle_jointure_renforcee(self, df):
        """Crée une clé de jointure renforcée avec hash"""
        try:
            # Combiner plusieurs colonnes pour une clé plus robuste
            colonnes_cles = []
            
            # Colonne principale (généralement l'ID de l'ordre)
            if len(df.columns) > 1:
                colonnes_cles.append(df.iloc[:, 1].astype(str))
            
            # Ajouter d'autres colonnes si disponibles
            if "Symbole" in df.columns:
                colonnes_cles.append(df["Symbole"].astype(str))
            if "Volume" in df.columns:
                colonnes_cles.append(df["Volume"].astype(str))
            if "Type" in df.columns:
                colonnes_cles.append(df["Type"].astype(str))
            
            # Créer une clé combinée
            if colonnes_cles:
                cle_combinee = "_".join(colonnes_cles)
                # Ajouter un hash pour éviter les collisions
                hash_cle = cle_combinee.apply(lambda x: hashlib.md5(str(x).encode()).hexdigest()[:8])
                return cle_combinee + "_" + hash_cle
            else:
                # Fallback sur la première colonne
                return df.iloc[:, 1].astype(str)
                
        except Exception as e:
            print(f"[WARNING] Error creating enhanced join key: {e}")
            # Fallback sur la méthode simple
            return df.iloc[:, 1].astype(str)
    
    def apply_matching_logic_renforcee(self, fusion_df):
        """Applique la logique de matching renforcée"""
        # Créer les clés pour les trades "in" avec plus de robustesse
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "in":
                # Essayer plusieurs méthodes de matching
                cle_match = None
                
                # Méthode 1: T/P ou S/L
                for val in ["T / P", "S / L"]:
                    if val in fusion_df.columns:
                        try:
                            prix = round(float(row[val]), 5)
                            if not pd.isna(prix):
                                cle_match = f"{row['Symbole_ordre']}-{prix}"
                                break
                        except:
                            continue
                
                # Méthode 2: Prix de transaction + volume
                if cle_match is None:
                    try:
                        prix = round(float(row["Prix_transaction"]), 5)
                        volume = str(row["Volume_ordre"])
                        cle_match = f"{row['Symbole_ordre']}-{prix}-{volume}"
                    except:
                        pass
                
                # Méthode 3: Timestamp + symbole
                if cle_match is None:
                    try:
                        timestamp = str(row.get("Heure d'ouverture", ""))
                        cle_match = f"{row['Symbole_ordre']}-{timestamp}"
                    except:
                        pass
                
                if cle_match:
                    fusion_df.at[idx, "Cle_Match"] = cle_match

        # Matcher les trades "out" avec plus de flexibilité
        all_match_values = set(fusion_df[fusion_df["Cle_Match"].notna()]["Cle_Match"].values)
        
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "out":
                # Essayer plusieurs méthodes de matching pour les sorties
                prix_match = self.extraire_prix_commentaire_renforcee(row)
                
                if prix_match:
                    # Essayer différentes combinaisons de clés
                    possibles_cles = [
                        f"{row['Symbole_ordre']}-{prix_match}",
                        f"{row['Symbole_ordre']}-{prix_match}-{row['Volume_ordre']}",
                        f"{row['Symbole_ordre']}-{prix_match}-{row.get('Heure d\'ouverture', '')}"
                    ]
                    
                    for cle_test in possibles_cles:
                        if cle_test in all_match_values:
                            fusion_df.at[idx, "Cle_Match"] = cle_test
                            break
    
    def extraire_prix_commentaire_renforcee(self, row):
        """Extrait le prix du commentaire avec plus de robustesse"""
        commentaire = str(row.get("Commentaire_ordre", ""))
        
        # Patterns multiples pour plus de robustesse
        patterns = [
            r'(tp|sl)[^\d]*(\d+[.,]?\d+)',
            r'(\d+[.,]?\d+)\s*(tp|sl)',
            r'(prix|price)[^\d]*(\d+[.,]?\d+)',
            r'(\d+[.,]?\d+)\s*(prix|price)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, commentaire.lower())
            if match:
                try:
                    prix = float(match.group(2).replace(",", "."))
                    return round(prix, 5)
                except:
                    continue
        
        return None
    
    def calculer_metriques_specialisees(self, df):
        """Calcule les métriques spécialisées par type d'instrument"""
        # Créer l'index des trades d'entrée
        df_in = df[(df["Direction"] == "in") & (df["Cle_Match"].notna())].copy()
        if len(df_in) > 0:
            df_in = df_in.set_index("Cle_Match")
        
        # Calculer les métriques selon le type d'instrument
        for idx, row in df.iterrows():
            type_inst = row.get("Type_Instrument", self.detecter_type_instrument(row["Symbole_ordre"]))
            
            if type_inst == InstrumentType.FOREX:
                df.at[idx, "Pips"] = self.calculer_pips_forex(row, df_in)
            elif type_inst == InstrumentType.METAUX:
                df.at[idx, "Points_Metaux"] = self.calculer_points_metaux(row, df_in)
            elif type_inst == InstrumentType.INDICES:
                df.at[idx, "Points_Indices"] = self.calculer_points_indices(row, df_in)
            elif type_inst == InstrumentType.CRYPTO:
                df.at[idx, "Points_Crypto"] = self.calculer_points_crypto(row, df_in)
            elif type_inst == InstrumentType.ENERGIE:
                df.at[idx, "Points_Energie"] = self.calculer_points_energie(row, df_in)
            else:
                df.at[idx, "Points_Actions"] = self.calculer_points_actions(row, df_in)
        
        return df
    
    def calculer_pips_forex(self, row, df_in):
        """Calcul des pips pour Forex"""
        symbole = str(row["Symbole_ordre"]).lower()
        profit = row["Profit"]
        volume_str = str(row["Volume_ordre"])
        
        if "/" in volume_str:
            volume = float(volume_str.split("/")[0].strip())
        else:
            volume = float(volume_str.strip())
        
        pip_size = 0.01 if "jpy" in symbole else 0.0001
        
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            pips = (prix_out - prix_in) / pip_size
                        else:
                            pips = (prix_in - prix_out) / pip_size
                        return round(pips, 2)
            
            # Fallback
            if "jpy" in symbole:
                valeur_pip = volume * pip_size * 1000
            else:
                valeur_pip = volume * pip_size * 100000
            
            if valeur_pip != 0:
                return round(profit / valeur_pip, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def calculer_points_metaux(self, row, df_in):
        """Calcul des points pour métaux précieux"""
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points = prix_out - prix_in
                        else:
                            points = prix_in - prix_out
                        return round(points, 2)
            
            # Fallback basé sur le profit
            volume = float(str(row["Volume_ordre"]).split("/")[0].strip())
            valeur_point = volume * 1.0  # ~1€ par point pour 1 lot
            
            if valeur_point != 0:
                return round(row["Profit"] / valeur_point, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def calculer_points_indices(self, row, df_in):
        """Calcul des points pour indices"""
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points = prix_out - prix_in
                        else:
                            points = prix_in - prix_out
                        return round(points, 2)
            
            # Fallback avec valeur variable selon l'indice
            symbole = str(row["Symbole_ordre"]).lower()
            volume = float(str(row["Volume_ordre"]).split("/")[0].strip())
            
            if "dax" in symbole or "ger30" in symbole:
                valeur_point = volume * 5.0
            elif "cac" in symbole or "fra40" in symbole:
                valeur_point = volume * 2.0
            elif "sp500" in symbole or "us500" in symbole:
                valeur_point = volume * 10.0
            else:
                valeur_point = volume * 5.0
            
            if valeur_point != 0:
                return round(row["Profit"] / valeur_point, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def calculer_points_crypto(self, row, df_in):
        """Calcul des points pour crypto"""
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points = prix_out - prix_in
                        else:
                            points = prix_in - prix_out
                        return round(points, 2)
            
            # Fallback
            volume = float(str(row["Volume_ordre"]).split("/")[0].strip())
            valeur_point = volume * 0.1  # Estimation pour crypto
            
            if valeur_point != 0:
                return round(row["Profit"] / valeur_point, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def calculer_points_energie(self, row, df_in):
        """Calcul des points pour énergie"""
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points = prix_out - prix_in
                        else:
                            points = prix_in - prix_out
                        return round(points, 2)
            
            # Fallback
            volume = float(str(row["Volume_ordre"]).split("/")[0].strip())
            valeur_point = volume * 10.0  # ~10€ par point pour 1 lot
            
            if valeur_point != 0:
                return round(row["Profit"] / valeur_point, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def calculer_points_actions(self, row, df_in):
        """Calcul des points pour actions"""
        try:
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            points = prix_out - prix_in
                        else:
                            points = prix_in - prix_out
                        return round(points, 2)
            
            # Fallback
            volume = float(str(row["Volume_ordre"]).split("/")[0].strip())
            valeur_point = volume * 1.0  # Estimation générale
            
            if valeur_point != 0:
                return round(row["Profit"] / valeur_point, 2)
            else:
                return None
                
        except Exception:
            return None 

    def trouver_ligne(self, df, mot_approx):
        """Trouve une ligne contenant un mot approximatif"""
        for i, row in df.iterrows():
            texte = row.astype(str).str.lower().str.replace(" ", "").str.replace(":", "")
            if texte.str.contains(mot_approx.lower()).any():
                return i
        raise ValueError(f"Ligne avec '{mot_approx}' non trouvée.")
    
    def safe_convert_to_float(self, series):
        """Convertit une série en float en gérant les valeurs NaN"""
        return pd.to_numeric(series.astype(str).str.replace(",", ".").replace("nan", ""), errors='coerce')
    
    def detecter_type_instrument(self, symbole):
        """Détecte le type d'instrument financier"""
        symbole = str(symbole).lower()
        
        # Vérifier d'abord si c'est du Forex
        if any(forex_pair in symbole for forex_pair in self.forex_pairs):
            return InstrumentType.FOREX
        
        # Vérifier les autres types
        for instrument_type, patterns in self.instrument_patterns.items():
            if any(pattern in symbole for pattern in patterns):
                return instrument_type
        
        # Par défaut
        return InstrumentType.ACTIONS
    
    def filtrer_par_instrument(self, df, instrument_filter):
        """Filtre le DataFrame selon le type d'instrument"""
        if instrument_filter is None:
            # Pas de filtre, garder tous les instruments
            df['Type_Instrument'] = df['Symbole_ordre'].apply(self.detecter_type_instrument)
            return df
        
        if instrument_filter == 'forex':
            # Garder uniquement le Forex
            df = df[df["Symbole_ordre"].apply(self.est_forex)]
            df['Type_Instrument'] = InstrumentType.FOREX
        elif instrument_filter == 'autres':
            # Exclure le Forex
            df = df[df["Symbole_ordre"].apply(self.est_autre_instrument)]
            df['Type_Instrument'] = df['Symbole_ordre'].apply(self.detecter_type_instrument)
        
        return df
    
    def est_forex(self, symbole):
        """Vérifie si un symbole est une paire Forex"""
        symbole = str(symbole).lower()
        return any(forex_pair in symbole for forex_pair in self.forex_pairs)
    
    def est_autre_instrument(self, symbole):
        """Vérifie si un symbole N'EST PAS une paire Forex"""
        return not self.est_forex(symbole)
    
    def fusionner_et_calculer_cumuls(self, tous_les_df):
        """Fusionne tous les DataFrames et calcule les intérêts composés + drawdown"""
        print(f"[DEBUG] Starting fusion and compound calculations...")
        
        # Fusionner tous les DataFrames
        df_complet = pd.concat(tous_les_df, ignore_index=True)
        print(f"[DEBUG] Merged {len(tous_les_df)} dataframes into {len(df_complet)} total trades")
        
        # Tri par date
        if "Heure d'ouverture" in df_complet.columns:
            df_complet["Date_parsed"] = pd.to_datetime(df_complet["Heure d'ouverture"], errors='coerce')
            df_complet = df_complet.sort_values("Date_parsed").reset_index(drop=True)
            df_complet = df_complet.drop("Date_parsed", axis=1)
        
        # Calculs cumulés avec intérêts composés
        df_complet["Profit_compose"] = 0.0
        df_complet["Profit_cumule"] = 0.0
        df_complet["Solde_cumule"] = 0.0
        df_complet["Drawdown_pct"] = 0.0
        df_complet["Drawdown_euros"] = 0.0
        df_complet["Drawdown_running_pct"] = 0.0
        
        solde_courant = self.solde_initial
        profit_cumule_reel = 0.0
        plus_haut_solde = self.solde_initial
        drawdown_running_max = 0.0
        
        print(f"[DEBUG] Starting compound interest calculations...")
        
        for idx, row in df_complet.iterrows():
            profit_original = row["Profit"] if pd.notna(row["Profit"]) else 0
            
            # Calculer le rendement en pourcentage
            if profit_original != 0 and self.solde_initial != 0:
                rendement_trade_pct = (profit_original / self.solde_initial) * 100
                profit_compose = (rendement_trade_pct / 100) * solde_courant
            else:
                profit_compose = 0
            
            # Mise à jour des cumuls
            solde_courant += profit_compose
            profit_cumule_reel += profit_compose
            
            # Mise à jour du plus haut solde historique
            if solde_courant > plus_haut_solde:
                plus_haut_solde = solde_courant
            
            # Calcul du drawdown CLASSIQUE
            if solde_courant < plus_haut_solde:
                drawdown_euros = plus_haut_solde - solde_courant
                drawdown_pct = (drawdown_euros / plus_haut_solde * 100)
            else:
                drawdown_euros = 0.0
                drawdown_pct = 0.0
            
            # Calcul du drawdown RUNNING (lissé)
            drawdown_actuel = (plus_haut_solde - solde_courant) / plus_haut_solde * 100
            if drawdown_actuel > drawdown_running_max:
                drawdown_running_max = drawdown_actuel
            
            if drawdown_actuel < drawdown_running_max:
                if profit_original > 0:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max * 0.9)
                else:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max)
            
            # Enregistrer les valeurs
            df_complet.at[idx, "Profit_compose"] = round(profit_compose, 2)
            df_complet.at[idx, "Profit_cumule"] = round(profit_cumule_reel, 2)
            df_complet.at[idx, "Solde_cumule"] = round(solde_courant, 2)
            df_complet.at[idx, "Drawdown_pct"] = round(drawdown_pct, 2)
            df_complet.at[idx, "Drawdown_euros"] = round(drawdown_euros, 2)
            df_complet.at[idx, "Drawdown_running_pct"] = round(drawdown_running_max, 2)
        
        print(f"[DEBUG] Compound calculations completed. Final solde: {solde_courant:.2f}")
        print(f"[DEBUG] Max drawdown: {df_complet['Drawdown_pct'].max():.2f}%")
        
        return df_complet
    
    def calculer_statistiques_avancees(self, df):
        """Calcule les statistiques avancées"""
        stats = {}
        
        # Séparation des trades par résultat (EXCLUANT LES NEUTRES)
        trades_gagnants = df[df["Profit"] > 0]["Profit"]
        trades_perdants = df[df["Profit"] < 0]["Profit"]
        
        # Moyennes
        stats["gain_moyen"] = trades_gagnants.mean() if len(trades_gagnants) > 0 else 0
        stats["perte_moyenne"] = trades_perdants.mean() if len(trades_perdants) > 0 else 0
        
        # Calcul des séries consécutives
        series_gagnantes = []
        series_perdantes = []
        
        serie_gagnante_actuelle = 0
        serie_perdante_actuelle = 0
        
        for _, row in df.iterrows():
            profit = row["Profit"]
            
            if profit > 0:
                serie_gagnante_actuelle += 1
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
            elif profit < 0:
                serie_perdante_actuelle += 1
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
            else:
                # Trade neutre
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
        
        # Ajouter la dernière série
        if serie_gagnante_actuelle > 0:
            series_gagnantes.append(serie_gagnante_actuelle)
        if serie_perdante_actuelle > 0:
            series_perdantes.append(serie_perdante_actuelle)
        
        stats["gains_consecutifs_max"] = max(series_gagnantes) if series_gagnantes else 0
        stats["pertes_consecutives_max"] = max(series_perdantes) if series_perdantes else 0
        
        # Statistiques du drawdown
        stats["drawdown_max_pct"] = df["Drawdown_pct"].max()
        stats["drawdown_max_euros"] = df["Drawdown_euros"].max()
        
        # Nombre de périodes de drawdown
        periodes_drawdown = len(df[df["Drawdown_pct"] > 0])
        stats["periodes_drawdown"] = periodes_drawdown
        
        return stats
    
    def create_excel_report(self, df_final, reports_folder, timestamp, instrument_filter=None):
        """Crée un rapport Excel complet avec feuilles spécialisées"""
        try:
            print(f"[DEBUG] Starting Excel report creation with specialized sheets")
            
            # Calculer les statistiques avancees
            stats_avancees = self.calculer_statistiques_avancees(df_final)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # === ONGLET 1: RÉSUMÉ GLOBAL ===
            ws_resume = wb.create_sheet("📊 Résumé Global")
            
            # Titre principal
            ws_resume.merge_cells('A1:H1')
            cell_titre = ws_resume['A1']
            cell_titre.value = f"RAPPORT TRADING AMÉLIORÉ - {datetime.now().strftime('%d/%m/%Y')}"
            cell_titre.font = Font(size=16, bold=True, color="FFFFFF")
            cell_titre.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_titre.alignment = Alignment(horizontal="center", vertical="center")
            
            # Statistiques globales
            total_trades = len(df_final)
            trades_gagnants = len(df_final[df_final["Profit"] > 0])
            trades_perdants = len(df_final[df_final["Profit"] < 0])
            trades_neutres = len(df_final[df_final["Profit"] == 0])
            trades_avec_resultat = trades_gagnants + trades_perdants
            
            profit_total_lineaire = df_final['Profit'].sum()
            profit_total_compose = df_final['Profit_cumule'].iloc[-1] if len(df_final) > 0 else 0
            solde_final = df_final['Solde_cumule'].iloc[-1] if len(df_final) > 0 else self.solde_initial
            rendement_pct = ((solde_final - self.solde_initial) / self.solde_initial * 100)
            
            difference_compose = profit_total_compose - profit_total_lineaire
            gain_compose_pct = ((profit_total_compose / profit_total_lineaire - 1) * 100) if profit_total_lineaire != 0 else 0
            taux_reussite = (trades_gagnants / trades_avec_resultat * 100) if trades_avec_resultat > 0 else 0
            
            # Tableau des statistiques globales
            stats_data = [
                ["📊 STATISTIQUES GLOBALES", ""],
                ["", ""],
                ["💰 Solde initial", f"{self.solde_initial:,.2f} €"],
                ["💳 Solde final (composé)", f"{solde_final:,.2f} €"],
                ["📈 Profit total (linéaire)", f"{profit_total_lineaire:,.2f} €"],
                ["🚀 Profit total (composé)", f"{profit_total_compose:,.2f} €"],
                ["⚡ Gain intérêts composés", f"{difference_compose:,.2f} € (+{gain_compose_pct:.2f}%)"],
                ["📊 Rendement global", f"{rendement_pct:.2f} %"],
                ["", ""],
                ["📉 ANALYSE DU DRAWDOWN", ""],
                ["", ""],
                ["📉 Drawdown maximum", f"{stats_avancees['drawdown_max_pct']:.2f} %"],
                ["💸 Drawdown max (euros)", f"{stats_avancees['drawdown_max_euros']:,.2f} €"],
                ["⏱️ Périodes de drawdown", f"{stats_avancees['periodes_drawdown']} trades"],
                ["", ""],
                ["🔢 ANALYSE DES TRADES (hors neutres)", ""],
                ["", ""],
                ["💱 Total trades", total_trades],
                ["✅ Trades gagnants", trades_gagnants],
                ["❌ Trades perdants", trades_perdants],
                ["⚪ Trades neutres (exclus)", f"{trades_neutres} (non comptés)"],
                ["🎯 Taux de réussite", f"{taux_reussite:.1f} % (sur {trades_avec_resultat} trades)"],
                ["", ""],
                ["📈 SÉRIES ET MOYENNES", ""],
                ["", ""],
                ["🔥 Gains consécutifs max", f"{stats_avancees['gains_consecutifs_max']} trades"],
                ["💔 Pertes consécutives max", f"{stats_avancees['pertes_consecutives_max']} trades"],
                ["💚 Gain moyen", f"{stats_avancees['gain_moyen']:,.2f} €"],
                ["💔 Perte moyenne", f"{stats_avancees['perte_moyenne']:,.2f} €"],
            ]
            
            # Ajouter les statistiques par type d'instrument si analyse complète
            if instrument_filter is None and 'Type_Instrument' in df_final.columns:
                stats_data.extend([
                    ["", ""],
                    ["📊 STATISTIQUES PAR TYPE D'INSTRUMENT", ""],
                    ["", ""]
                ])
                
                for inst_type in InstrumentType:
                    df_type = df_final[df_final['Type_Instrument'] == inst_type]
                    if len(df_type) > 0:
                        profit_type = df_type['Profit'].sum()
                        nb_trades = len(df_type)
                        taux_type = len(df_type[df_type['Profit'] > 0]) / nb_trades * 100 if nb_trades > 0 else 0
                        
                        stats_data.extend([
                            [f"🏷️ {str(inst_type.value).upper()}", ""],
                            [f"  - Nombre de trades", nb_trades],
                            [f"  - Profit total", f"{profit_type:,.2f} €"],
                            [f"  - Taux de réussite", f"{taux_type:.1f} %"],
                            ["", ""]
                        ])
            
            # Ajout des statistiques par fichier si disponibles
            if self.statistiques_fichiers:
                stats_data.extend([
                    ["", ""],
                    ["📁 DÉTAIL PAR FICHIER", ""],
                    ["", ""]
                ])
                for fichier, stats in self.statistiques_fichiers.items():
                    stats_data.append([f"📄 {fichier[:30]}...", f"{stats['trades']} trades, {stats['exclus']} exclus"])
            
            for row_idx, (label, value) in enumerate(stats_data, 3):
                ws_resume[f'A{row_idx}'] = label
                ws_resume[f'B{row_idx}'] = value
                
                # Formatage des en-têtes
                if any(word in label for word in ["STATISTIQUES", "ANALYSE", "DRAWDOWN", "SÉRIES", "DÉTAIL"]):
                    ws_resume[f'A{row_idx}'].font = Font(bold=True, color="366092")
                    ws_resume[f'A{row_idx}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                ws_resume[f'A{row_idx}'].alignment = Alignment(horizontal="left")
                ws_resume[f'B{row_idx}'].alignment = Alignment(horizontal="right")
            
            print(f"[DEBUG] Summary sheet created with {len(stats_data)} rows")
            
            # === ONGLET 2: DONNÉES BRUTES UNIFIÉES ===
            ws_data = wb.create_sheet("📋 Données Unifiées")
            
            # Colonnes unifiées (sans pips/points spécifiques)
            colonnes_unifiees = [
                "Heure d'ouverture", "Ordre_ordre", "Symbole_ordre", "Type_ordre", 
                "Volume_ordre", "S / L", "T / P", "Direction", "Prix_transaction",
                "Profit", "Cle_Match", "Type_Instrument", "Profit_compose", 
                "Profit_cumule", "Solde_cumule", "Drawdown_pct", "Drawdown_euros"
            ]
            
            colonnes_finales = [col for col in colonnes_unifiees if col in df_final.columns]
            df_unifie = df_final[colonnes_finales]
            
            # Insérer les données unifiées
            for r in dataframe_to_rows(df_unifie, index=False, header=True):
                ws_data.append(r)
            
            # Formatage des en-têtes
            for cell in ws_data[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            print(f"[DEBUG] Unified data sheet created with {len(df_unifie)} rows")
            
            # === ONGLETS SPÉCIALISÉS PAR TYPE ===
            if 'Type_Instrument' in df_final.columns:
                self.creer_onglets_specialises(wb, df_final)
            
            # Ajuster la largeur des colonnes
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            fichier_rapport = os.path.join(reports_folder, f"RAPPORT_AMELIORE_{timestamp}.xlsx")
            wb.save(fichier_rapport)
            
            print(f"[DEBUG] Excel report saved successfully: {fichier_rapport}")
            return fichier_rapport
            
        except Exception as e:
            print(f"[ERROR] Error creating Excel report: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors de la création du rapport Excel: {str(e)}")
    
    def creer_onglets_specialises(self, wb, df_final):
        """Crée des onglets spécialisés pour chaque type d'instrument"""
        
        # === ONGLET FOREX ===
        df_forex = df_final[df_final['Type_Instrument'] == InstrumentType.FOREX]
        if len(df_forex) > 0:
            ws_forex = wb.create_sheet("💱 Analyse Forex")
            self.creer_onglet_forex(ws_forex, df_forex)
        
        # === ONGLET MÉTAUX ===
        df_metaux = df_final[df_final['Type_Instrument'] == InstrumentType.METAUX]
        if len(df_metaux) > 0:
            ws_metaux = wb.create_sheet("🥇 Analyse Métaux")
            self.creer_onglet_metaux(ws_metaux, df_metaux)
        
        # === ONGLET INDICES ===
        df_indices = df_final[df_final['Type_Instrument'] == InstrumentType.INDICES]
        if len(df_indices) > 0:
            ws_indices = wb.create_sheet("📊 Analyse Indices")
            self.creer_onglet_indices(ws_indices, df_indices)
        
        # === ONGLET CRYPTO ===
        df_crypto = df_final[df_final['Type_Instrument'] == InstrumentType.CRYPTO]
        if len(df_crypto) > 0:
            ws_crypto = wb.create_sheet("₿ Analyse Crypto")
            self.creer_onglet_crypto(ws_crypto, df_crypto)
        
        # === ONGLET ÉNERGIE ===
        df_energie = df_final[df_final['Type_Instrument'] == InstrumentType.ENERGIE]
        if len(df_energie) > 0:
            ws_energie = wb.create_sheet("🛢️ Analyse Énergie")
            self.creer_onglet_energie(ws_energie, df_energie)
        
        # === ONGLET ACTIONS ===
        df_actions = df_final[df_final['Type_Instrument'] == InstrumentType.ACTIONS]
        if len(df_actions) > 0:
            ws_actions = wb.create_sheet("📈 Analyse Actions")
            self.creer_onglet_actions(ws_actions, df_actions)
    
    def creer_onglet_forex(self, ws, df):
        """Crée l'onglet spécialisé Forex avec pips"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE FOREX - CALCUL DES PIPS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec pips
        colonnes_forex = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Pips"
        ]
        
        colonnes_disponibles = [col for col in colonnes_forex if col in df.columns]
        df_forex = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_forex.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Forex
        if 'Pips' in df.columns:
            pips_total = df['Pips'].sum()
            pips_moyen = df['Pips'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES FOREX"
            ws['A' + str(len(df) + 7)] = f"Pips totaux: {pips_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Pips moyen: {pips_moyen:.2f}"
    
    def creer_onglet_metaux(self, ws, df):
        """Crée l'onglet spécialisé Métaux avec points"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE MÉTAUX - CALCUL DES POINTS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec points métaux
        colonnes_metaux = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Points_Metaux"
        ]
        
        colonnes_disponibles = [col for col in colonnes_metaux if col in df.columns]
        df_metaux = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_metaux.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Métaux
        if 'Points_Metaux' in df.columns:
            points_total = df['Points_Metaux'].sum()
            points_moyen = df['Points_Metaux'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES MÉTAUX"
            ws['A' + str(len(df) + 7)] = f"Points totaux: {points_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Points moyen: {points_moyen:.2f}"
    
    def creer_onglet_indices(self, ws, df):
        """Crée l'onglet spécialisé Indices avec points"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE INDICES - CALCUL DES POINTS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec points indices
        colonnes_indices = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Points_Indices"
        ]
        
        colonnes_disponibles = [col for col in colonnes_indices if col in df.columns]
        df_indices = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_indices.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Indices
        if 'Points_Indices' in df.columns:
            points_total = df['Points_Indices'].sum()
            points_moyen = df['Points_Indices'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES INDICES"
            ws['A' + str(len(df) + 7)] = f"Points totaux: {points_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Points moyen: {points_moyen:.2f}"
    
    def creer_onglet_crypto(self, ws, df):
        """Crée l'onglet spécialisé Crypto avec points"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE CRYPTO - CALCUL DES POINTS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec points crypto
        colonnes_crypto = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Points_Crypto"
        ]
        
        colonnes_disponibles = [col for col in colonnes_crypto if col in df.columns]
        df_crypto = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_crypto.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Crypto
        if 'Points_Crypto' in df.columns:
            points_total = df['Points_Crypto'].sum()
            points_moyen = df['Points_Crypto'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES CRYPTO"
            ws['A' + str(len(df) + 7)] = f"Points totaux: {points_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Points moyen: {points_moyen:.2f}"
    
    def creer_onglet_energie(self, ws, df):
        """Crée l'onglet spécialisé Énergie avec points"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE ÉNERGIE - CALCUL DES POINTS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec points énergie
        colonnes_energie = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Points_Energie"
        ]
        
        colonnes_disponibles = [col for col in colonnes_energie if col in df.columns]
        df_energie = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_energie.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Énergie
        if 'Points_Energie' in df.columns:
            points_total = df['Points_Energie'].sum()
            points_moyen = df['Points_Energie'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES ÉNERGIE"
            ws['A' + str(len(df) + 7)] = f"Points totaux: {points_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Points moyen: {points_moyen:.2f}"
    
    def creer_onglet_actions(self, ws, df):
        """Crée l'onglet spécialisé Actions avec points"""
        # Titre
        ws.merge_cells('A1:H1')
        cell_titre = ws['A1']
        cell_titre.value = "ANALYSE ACTIONS - CALCUL DES POINTS"
        cell_titre.font = Font(size=14, bold=True, color="FFFFFF")
        cell_titre.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        cell_titre.alignment = Alignment(horizontal="center")
        
        # Données avec points actions
        colonnes_actions = [
            "Heure d'ouverture", "Symbole_ordre", "Type_ordre", "Volume_ordre", 
            "Direction", "Prix_transaction", "Profit", "Points_Actions"
        ]
        
        colonnes_disponibles = [col for col in colonnes_actions if col in df.columns]
        df_actions = df[colonnes_disponibles]
        
        # En-têtes
        for col_idx, header in enumerate(colonnes_disponibles, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_idx, (_, row) in enumerate(df_actions.iterrows(), 4):
            for col_idx, col in enumerate(colonnes_disponibles, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
        
        # Statistiques Actions
        if 'Points_Actions' in df.columns:
            points_total = df['Points_Actions'].sum()
            points_moyen = df['Points_Actions'].mean()
            
            ws['A' + str(len(df) + 6)] = "📊 STATISTIQUES ACTIONS"
            ws['A' + str(len(df) + 7)] = f"Points totaux: {points_total:.2f}"
            ws['A' + str(len(df) + 8)] = f"Points moyen: {points_moyen:.2f}"

# Exemple d'utilisation
if __name__ == "__main__":
    # Créer l'analyseur amélioré
    analyzer = TradingAnalyzerImproved(solde_initial=10000)
    
    # Exemple d'utilisation pour tous les instruments
    # df_complet = analyzer.process_files(file_paths, task_id, task_status, None)
    # analyzer.create_excel_report(df_complet, reports_folder, timestamp, None) 