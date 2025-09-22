#!/usr/bin/env python3
"""
Analyseur de Trading Simplifié - Sans Pandas
Version optimisée pour Render avec uniquement openpyxl
"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference, PieChart, BarChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TradingAnalyzerSimple:
    def __init__(self, solde_initial=10000):
        self.solde_initial = solde_initial
        self.statistiques_fichiers = {}
        
        # Configuration des symboles par type
        self.symboles_forex = [
            "eurusd", "gbpusd", "usdchf", "usdjpy", "usdcad", "audusd", "nzdusd",
            "eurjpy", "gbpjpy", "audjpy", "cadjpy", "chfjpy", "nzdjpy",
            "eurgbp", "euraud", "eurcad", "eurchf", "eurnzd",
            "gbpaud", "gbpcad", "gbpchf", "gbpnzd",
            "audcad", "audchf", "audnzd", "cadchf", "nzdcad", "nzdchf"
        ]
        
        self.symboles_autres = ["gold", "xauusd", "xag", "silver", "dax", "cac", "sp500", "btc", "eth", "oil", "wti"]

    def process_files(self, file_paths, task_id, task_status, filter_type=None):
        """Traite une liste de fichiers Excel"""
        try:
            print(f"[DEBUG] Starting analysis with {len(file_paths)} files, filter: {filter_type}")
            tous_les_resultats = []
            total_files = len(file_paths)
            
            for i, file_path in enumerate(file_paths):
                progress = 20 + (i / total_files) * 40
                task_status[task_id]['progress'] = int(progress)
                task_status[task_id]['message'] = f'Traitement du fichier {i+1}/{total_files}...'
                
                print(f"[DEBUG] Processing file {i+1}/{total_files}: {os.path.basename(file_path)}")
                
                trades = self.process_single_file(file_path, filter_type)
                
                if trades:
                    tous_les_resultats.extend(trades)
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': len(trades),
                        'exclus': 0,
                        'doublons': 0,
                        'erreur': None
                    }
                    print(f"[DEBUG] File processed successfully: {len(trades)} trades")
                else:
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': 0,
                        'exclus': 0,
                        'doublons': 0,
                        'erreur': "Aucune donnée trouvée"
                    }
            
            if not tous_les_resultats:
                print(f"[DEBUG] No valid data found in any file")
                return None
            
            task_status[task_id]['progress'] = 60
            task_status[task_id]['message'] = 'Calculs des intérêts composés...'
            
            # Calculs cumulés
            final_data = self.calculer_cumuls(tous_les_resultats)
            
            task_status[task_id]['progress'] = 75
            task_status[task_id]['message'] = 'Calculs des statistiques...'
            
            return final_data
            
        except Exception as e:
            print(f"[ERROR] Error in process_files: {str(e)}")
            raise Exception(f"Erreur lors du traitement des fichiers: {str(e)}")

    def process_single_file(self, file_path, filter_type=None):
        """Traite un seul fichier Excel"""
        try:
            print(f"[DEBUG] Opening file: {file_path}")
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Collecter toutes les données
            trades = []
            
            # Parcourir toutes les lignes pour trouver les trades
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and len(row) >= 8:  # Minimum de colonnes attendues
                    try:
                        # Essayer de détecter une ligne de trade
                        symbole = str(row[2] or "").strip().lower() if len(row) > 2 else ""
                        profit_str = str(row[6] or "") if len(row) > 6 else ""
                        
                        if symbole and profit_str:
                            # Convertir le profit
                            profit = self.safe_float(profit_str)
                            if profit is not None:
                                # Appliquer le filtre
                                if self.should_include_trade(symbole, filter_type):
                                    trade = {
                                        'date': str(row[0] or ""),
                                        'ordre': str(row[1] or ""),
                                        'symbole': symbole,
                                        'type': str(row[3] or ""),
                                        'volume': self.safe_float(str(row[4] or "")),
                                        'prix': self.safe_float(str(row[5] or "")),
                                        'profit': profit,
                                        'commentaire': str(row[7] or "") if len(row) > 7 else ""
                                    }
                                    trades.append(trade)
                    except Exception as e:
                        continue  # Ignorer les lignes problématiques
            
            print(f"[DEBUG] Found {len(trades)} trades in file")
            return trades
            
        except Exception as e:
            print(f"[ERROR] Error processing file {file_path}: {str(e)}")
            return []

    def safe_float(self, value):
        """Convertit une valeur en float de manière sécurisée"""
        if not value:
            return 0.0
        try:
            # Nettoyer la valeur
            clean_value = str(value).replace(",", ".").replace(" ", "")
            clean_value = re.sub(r'[^\d.-]', '', clean_value)
            return float(clean_value) if clean_value else 0.0
        except:
            return 0.0

    def should_include_trade(self, symbole, filter_type):
        """Détermine si un trade doit être inclus selon le filtre"""
        if not filter_type:
            return True
        
        symbole = symbole.lower()
        
        if filter_type == 'forex':
            return any(forex in symbole for forex in self.symboles_forex)
        elif filter_type == 'autres':
            return not any(forex in symbole for forex in self.symboles_forex)
        
        return True

    def calculer_cumuls(self, trades):
        """Calcule les cumuls et intérêts composés"""
        print(f"[DEBUG] Calculating compound interest for {len(trades)} trades")
        
        # Trier par date si possible
        try:
            trades.sort(key=lambda x: x.get('date', ''))
        except:
            pass
        
        solde_courant = self.solde_initial
        profit_cumule = 0.0
        plus_haut_solde = self.solde_initial
        
        for trade in trades:
            profit = trade.get('profit', 0)
            
            # Calcul intérêts composés
            if profit != 0 and self.solde_initial != 0:
                rendement_pct = (profit / self.solde_initial) * 100
                profit_compose = (rendement_pct / 100) * solde_courant
            else:
                profit_compose = 0
            
            solde_courant += profit_compose
            profit_cumule += profit_compose
            
            if solde_courant > plus_haut_solde:
                plus_haut_solde = solde_courant
            
            # Calcul drawdown
            drawdown_pct = 0.0
            if solde_courant < plus_haut_solde:
                drawdown_pct = ((plus_haut_solde - solde_courant) / plus_haut_solde) * 100
            
            # Ajouter les calculs au trade
            trade['profit_compose'] = round(profit_compose, 2)
            trade['profit_cumule'] = round(profit_cumule, 2)
            trade['solde_cumule'] = round(solde_courant, 2)
            trade['drawdown_pct'] = round(drawdown_pct, 2)
        
        print(f"[DEBUG] Compound calculations completed. Final balance: {solde_courant:.2f}")
        return trades

    def create_excel_report(self, trades, reports_folder, timestamp, filter_type=None):
        """Crée un rapport Excel simple"""
        try:
            print(f"[DEBUG] Creating Excel report with {len(trades)} trades")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapport Trading"
            
            # En-têtes
            headers = [
                "Date", "Ordre", "Symbole", "Type", "Volume", "Prix", 
                "Profit", "Profit Composé", "Profit Cumulé", "Solde Cumulé", "Drawdown %"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Données
            for row_idx, trade in enumerate(trades, 2):
                ws.cell(row=row_idx, column=1, value=trade.get('date', ''))
                ws.cell(row=row_idx, column=2, value=trade.get('ordre', ''))
                ws.cell(row=row_idx, column=3, value=trade.get('symbole', ''))
                ws.cell(row=row_idx, column=4, value=trade.get('type', ''))
                ws.cell(row=row_idx, column=5, value=trade.get('volume', 0))
                ws.cell(row=row_idx, column=6, value=trade.get('prix', 0))
                ws.cell(row=row_idx, column=7, value=trade.get('profit', 0))
                ws.cell(row=row_idx, column=8, value=trade.get('profit_compose', 0))
                ws.cell(row=row_idx, column=9, value=trade.get('profit_cumule', 0))
                ws.cell(row=row_idx, column=10, value=trade.get('solde_cumule', 0))
                ws.cell(row=row_idx, column=11, value=trade.get('drawdown_pct', 0))
            
            # Ajuster les largeurs de colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            suffix = f"_{filter_type.upper()}" if filter_type else "_UNIFIED"
            fichier_rapport = os.path.join(reports_folder, f"RAPPORT_TRADING{suffix}_{timestamp}.xlsx")
            wb.save(fichier_rapport)
            
            print(f"[DEBUG] Excel report saved: {fichier_rapport}")
            return fichier_rapport
            
        except Exception as e:
            print(f"[ERROR] Error creating Excel report: {str(e)}")
            raise Exception(f"Erreur lors de la création du rapport Excel: {str(e)}")

    def calculer_statistiques_avancees(self, trades):
        """Calcule les statistiques avancées"""
        if not trades:
            return {}
        
        profits = [trade.get('profit', 0) for trade in trades]
        trades_gagnants = [p for p in profits if p > 0]
        trades_perdants = [p for p in profits if p < 0]
        
        stats = {
            'total_trades': len(trades),
            'trades_gagnants': len(trades_gagnants),
            'trades_perdants': len(trades_perdants),
            'profit_total': sum(profits),
            'gain_moyen': sum(trades_gagnants) / len(trades_gagnants) if trades_gagnants else 0,
            'perte_moyenne': sum(trades_perdants) / len(trades_perdants) if trades_perdants else 0,
            'taux_reussite': (len(trades_gagnants) / len(trades) * 100) if trades else 0,
            'drawdown_max': max([trade.get('drawdown_pct', 0) for trade in trades]) if trades else 0
        }
        
        return stats